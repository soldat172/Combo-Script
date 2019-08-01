import keyboard, openpyxl, pyautogui, tkinter, time, sys, os, datetime
from tkinter import filedialog, messagebox, Tk
from openpyxl import Workbook, load_workbook
from guizero import App, Text, TextBox, PushButton, Window



global siteName
global apPrefix
global apNumber
global apName
global excelCount
global excelNumber
global rowLetter
global rowNumber
global rowWholeName
global list
global excelName
global formatNumber
global excelStartingInteger
global keyPress

siteName = str ('Paul')
apPrefix = siteName + '-XX-'
apNumber =  int('1')
apName = str(apPrefix) + str(apNumber)
excelCount = int('1')
excelNumber = int('1')
rowLetter = 'A'
rowNumber = int('1')

rowWholeName = str(rowLetter) + str(rowNumber)
list=[99999]
excelName = 'Paul'
formatNumber = format(apNumber, '05')
excelStartingInteger = int('1')
currentTime = datetime.datetime.now().strftime("%Y-%m-%d  %H.%M.%S")


def startWorkbook ():#Starts an excel. Required for visioTool.
    global wb
    wb = Workbook()
    global ws
    ws = wb.active

def openFiles():#Opens file explorer
    global excelName
    #Tk().withdraw() # We don't want a full GUI, so keep the root window from appearing
    excelName = filedialog.askopenfilename() # Show an "Open" dialog box and return the path to the selected file
    print(excelName)
    print('')

def pause(): #Pause logic
    keyPress = keyboard.read_key()
    if keyPress == 'pause': #Reads if pause button has been pressed and unpauses
        print('Unpaused')
        time.sleep(.3)
        exit
    else: #If pause has not been pressed then loop repeats infinitely WOO
        pause()
        
def theGUI(): #all the GUI stuff
    print('')
    print('Start of Visio Tool')
    def apStartingNumber(): # part of GUI that allows changing of the starting AP number
        global apNumber
        global formatNumber
        apNumber = int(startingNumber.value)
        formatNumber = format(apNumber, '05')
        print ('AP number changed to ' + str(apNumber))
        changingText.value = "AP number changed to " + str(apNumber) #Text for changing AP number

    def changeSiteName(): # part of GUI that allows changing of the site name
        global siteName
        global apPrefix
        global apName
        siteName = str(siteNames.value)
        apPrefix = siteName + '-XX-'
        apName = str(apPrefix) + str(apNumber)
        print('Site name changed to ' + siteName)
        changingText.value = "Site changed to " + siteName #Text for changing site name
        
    def directions(): #Once "Go?" has been pressed execute the following items
        startWorkbook()
        openFiles()
        app.hide()
        visioLoop()
        python = sys.executable #Restarts the whole program
        os.execl(python, python, * sys.argv) #Restarts the whole program
        
    app = App(title = "Phoenix_Oath", width=352, height=132, layout='grid')

    button7 = PushButton(app, text = "Go?", command = directions, grid=[2,3])

    #Logic for changing starting AP number
    startingNumberText = Text(app, text="AP Number?", align="left", grid=[0,1]) #Text asks for AP number
    startingNumber = TextBox(app, align="right",text = "1", width=30, grid=[1,1]) #Text box for data entry
    button4 = PushButton(app, text = "Confirm", command = apStartingNumber, grid=[2,1])

    #Logic for changing site name
    siteNamesText = Text(app, text="Site Name?", align="left",  grid=[0,2]) #Text asks for site name
    siteNames = TextBox(app, align="right",text = "PAUL", width=30, grid=[1,2]) #Text box for data entry
    button6 = PushButton(app, text = "Confirm", command = changeSiteName, grid=[2,2])

    changingText = Text(app,text="War has changed", align ="left", grid=[1,3])
    app.display() # initiates the GUI. Allowing it to be used
    
def saveExcel(): # saves Visio Tool names to Excel
    global displayText
    global excelName
    wb.save(excelName) # Saves workbook
    print ('')
    print ('Worksheet saved')
    print ('End of Visio Tool')
        
def visioGuts(): # the internals to the Visio Tool. Determines how most of the program is run
    global apNumber
    global formatNumber
    global excelStartingInteger
    print (apName)
    ws.cell(excelStartingInteger, 1, apName)  #writes in excel **format** ->(row, column, content to be written in cell)
    pyautogui.press('backspace'); pyautogui.typewrite(str(formatNumber)) # takes control of keyboard. hits backspace and types AP number
    apNumber += 1 #increments ap number up by 1
    excelStartingInteger += 1 # Increments Excel cell to be written in
    formatNumber = format(apNumber, '05') # Modifies apNumber by adding up to 5 zeros in front

def visioLoop():
    global apNumber
    global formatNumber
    global excelStartingInteger
    global apName
    while True:
        keyPress = keyboard.read_key() #reads ALL keypress's and saves to variable
        
        if keyPress == '`' : #Adds a Default Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber)
            visioGuts ()
            
        elif keyPress == 's': #Adds a Standup Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber) + 'S'
            visioGuts ()
            
        elif keyPress == 'g' : #Adds a Guard Shack Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber) + 'G'
            visioGuts ()
            
        elif keyPress == 'm' : #Adds a MOD/KIVA Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber) + 'M'
            visioGuts ()
            
        elif keyPress == 'h' : #Adds a High Racking Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber) + 'H'
            visioGuts ()
            
        elif keyPress == 'd' : #Adds a Door Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber) + 'D'
            visioGuts ()

        elif keyPress == 'e' : #Adds a External Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber) + 'E'
            visioGuts ()
            
        elif keyPress == '[': #goes down one ap number.
            apNumber -= 1
            excelStartingInteger -= 1
            formatNumber = format(apNumber, '05')
            apName = str(apPrefix) + str(formatNumber)
            print (apName + " fix?")
            time.sleep(.09)

        elif keyPress == ']': #goes up one ap number.
            apNumber += 1
            excelStartingInteger += 1
            formatNumber = format(apNumber, '05')
            apName = str(apPrefix) + str(formatNumber)
            print (apName + " fix?")
            time.sleep(.09)
            
        elif keyPress == 'pause': #Well it pauses everthing.....
            print('Paused')
            time.sleep(.3)
            pause() #Program is stuck in the pause loop until pause is pressed again

        elif keyPress == '=': #Saves the CAD Cutsheet and ends the Visio Tool   
            saveExcel()
            sys.exit()



class Options:
    def open_audit_sheet (self): #Opens a file explorer and returns path from chosen file
        global audit_sheet
        Tk().withdraw() # We don't want a full GUI, so keep the root window from appearing
        xlname = filedialog.askopenfilename() # Show an "Open" dialog box and return the path to the selected file
        print("Loading Excell File... ")
        audit_book = load_workbook(xlname)#Opens Excel
        audit_sheet = audit_book['Audit'] #grabs information from the "audit" sheet

    def audit_sheet_sorter (self): #reads all AP's from chosen Audit Sheet and categories them by their color
        global AP_red_list
        global AP_blue_list
        global AP_green_list
        global AP_orange_list
        
        xlname_col = 'A' #column letter to read AP names from
        xlname_col_fail = 'K' #column letter to read fails from
        xlname_row = int('3') #Starting row in excel
        xlname_row_fail = int('3') #Starting row in excel
    
        AP_red_list = []  #initiates a list for green AP's ('2')
        AP_blue_list = []  #initiates a list for green AP's ('4')
        AP_green_list = []  #initiates a list for green AP's ('1')
        AP_orange_list = []  #initiates a list for green AP's ('3')
        
        AP_red_count_list = int('0')#Keeps track how long the list is
        AP_blue_count_list = int('0')
        AP_green_count_list = int('0')
        AP_orange_count_list = int('0') 
        AP_count_total = int('0')          
                
        print("List of AP ID's:")
        for AP in range (0, audit_sheet.max_row +1): #range of APS being colored
            xlname_col_fail_num = str(xlname_col_fail) + str(xlname_row_fail)
            xlname_col_num = str(xlname_col) + str(xlname_row)
            cell_value1 = audit_sheet[xlname_col_fail_num]
            cell_value2 = audit_sheet[xlname_col_num]
            AP_fail = cell_value1.value
            AP_ID = cell_value2.value
            
            if AP_fail == 1: #reads all green AP's
                AP_green_list.append(AP_ID)
                print(AP_green_list[AP_green_count_list])
                AP_green_count_list += 1
                AP_count_total += 1

            elif AP_fail == 2: #reads all red AP's
                AP_red_list.append(AP_ID)
                print(AP_red_list[AP_red_count_list])
                AP_red_count_list += 1
                AP_count_total += 1
                
            elif AP_fail == 3: #reads all orange AP's
                AP_orange_list.append(AP_ID)
                print(AP_orange_list[AP_orange_count_list])
                AP_orange_count_list += 1
                AP_count_total += 1
                
            elif AP_fail == 4: #reads all blue AP's
                AP_blue_list.append(AP_ID)
                print(AP_blue_list[AP_blue_count_list])
                AP_blue_count_list += 1
                AP_count_total += 1

            elif AP_fail == 5: #reads all red and orange AP's
                AP_red_list.append(AP_ID)
                print(AP_red_list[AP_red_count_list])
                AP_red_count_list += 1
                AP_orange_list.append(AP_ID)
                print(AP_orange_list[AP_orange_count_list])
                AP_orange_count_list += 1
                AP_count_total += 1

            elif AP_fail == 6: #reads all red and blue AP's
                AP_red_list.append(AP_ID)
                print(AP_red_list[AP_red_count_list])
                AP_red_count_list += 1
                AP_blue_list.append(AP_ID)
                print(AP_blue_list[AP_blue_count_list])
                AP_blue_count_list += 1
                AP_count_total += 1

            elif AP_fail == 7: #reads all orange and blue AP's
                AP_orange_list.append(AP_ID)
                print(AP_orange_list[AP_orange_count_list])
                AP_orange_count_list += 1
                AP_blue_list.append(AP_ID)
                print(AP_blue_list[AP_blue_count_list])
                AP_blue_count_list += 1
                AP_count_total += 1

            elif AP_fail == 8: #reads all red and orange AP's
                AP_red_list.append(AP_ID)
                print(AP_red_list[AP_red_count_list])
                AP_red_count_list += 1
                AP_orange_list.append(AP_ID)
                print(AP_orange_list[AP_orange_count_list])
                AP_orange_count_list += 1    
                AP_blue_list.append(AP_ID)
                print(AP_blue_list[AP_blue_count_list])
                AP_blue_count_list += 1
                AP_count_total += 1

            else:
                print('End of Audit Sheets')
                break   
            xlname_row_fail += 1
            xlname_row += 1
        print(f'Surveyed APs: {AP_count_total}')
        print(f'Passed (Green) APs: {AP_green_count_list} ')
        print(f'Failed APs: {AP_count_total - AP_green_count_list}')
        print(f'Red APs: {AP_red_count_list} ')
        print(f'Blue APs: {AP_blue_count_list} ')
        print(f'Orange APs: {AP_orange_count_list}')

    def search_options (self):  #sets visio to search all pages for AP IDs, must be done before coloring
        print ("\nStarting Search...")
        pyautogui.hotkey('ctrl', 'f')
        pyautogui.press(['tab','tab'])
        pyautogui.press('down')
        pyautogui.press('esc')
        

    #def font_white(self): 
        #time.sleep(.05)
        #pyautogui.hotkey('alt', 'h')
        #pyautogui.hotkey('f', 'c')
        #pyautogui.press(['down','left','left','left','left','left','enter'])
        #print("Font set to white")

    def auto_script_grey(self):
        # cntr + page down/up to navigate visio pages
        pyautogui.hotkey('ctrl','a')
        pyautogui.hotkey('alt', 'h')#line coloring
        pyautogui.press(['l','down','down','down','down','down','left','left','left','enter'])
        pyautogui.hotkey('alt', 'h')#fill coloring
        pyautogui.press(['i','down','down','down','down','down','left','left','left','enter'])
        print("Set all APs to GREY\n")

    def auto_script_green(self):  #finds AP name in Visio and colors it GREEN,
        Options.search_options(self)
        AP_green_count = 0
        for x in AP_green_list: 
            AP_num = AP_green_list[AP_green_count]
            AP_num_print = (AP_num[8:13]) if len(AP_num) > 5 else AP_num #Truncates site name, IDF, and AP letter.
            print(AP_num_print)
            pyautogui.hotkey('ctrl', 'f')
            pyautogui.typewrite(AP_num_print)
            pyautogui.press(['enter', 'esc','esc','tab','enter'])
            pyautogui.hotkey('shift','tab','enter')
            pyautogui.hotkey('shift','tab','enter')
            pyautogui.hotkey('alt', 'h')#line coloring
            pyautogui.press(['l','down','down','down','down','down','down','down','enter'])
            pyautogui.hotkey('alt', 'h')#fill coloring
            pyautogui.press(['i','down','down','down','down','down','down','down','enter'])
            AP_green_count += 1     #adds one to total grey count
        print(f'Colored {AP_green_count} green APs')
        print("Select another color, or press 'QUIT' to end the program.\n")

    def auto_script_red(self): #finds AP name in Visio and colors it RED, 
        Options.search_options(self)    
        AP_red_count = 0
        for x in AP_red_list: #finds AP name in Visio and colors it
            AP_num = AP_red_list[AP_red_count]
            AP_num_print = (AP_num[8:13]) if len(AP_num) > 5 else AP_num #Truncates site name, IDF, and AP letter.
            print(AP_num_print)
            pyautogui.hotkey('ctrl', 'f')
            pyautogui.typewrite(AP_num_print)
            pyautogui.press(['enter', 'esc','esc','tab','enter'])
            pyautogui.hotkey('shift','tab','enter')
            pyautogui.hotkey('shift','tab','enter')
            pyautogui.hotkey('alt', 'h')#line coloring
            pyautogui.press(['l','down','down','down','down','down','down','down','left','left','left','left','enter'])
            pyautogui.hotkey('alt', 'h')#fill coloring
            pyautogui.press(['i','down','down','down','down','down','down','down','left','left','left','left','enter'])
            #Options.font_white(self)
            AP_red_count += 1   #adds one to total red count
        print(f'Colored {AP_red_count} red APs')
        print("Select another color, or press 'QUIT' to end the program.\n")

    def auto_script_orange(self): #finds AP name in Visio and colors it ORANGE,
        Options.search_options(self)
        AP_orange_count = 0
        for x in AP_orange_list: #finds AP name in Visio and colors it
            AP_num = AP_orange_list[AP_orange_count]
            AP_num_print = (AP_num[8:13]) if len(AP_num) > 5 else AP_num #Truncates site name, IDF, and AP letter.
            print(AP_num_print)
            pyautogui.hotkey('ctrl', 'f')
            pyautogui.typewrite(AP_num_print)
            pyautogui.press(['enter', 'esc','esc','tab','enter'])
            pyautogui.hotkey('shift','tab','enter')
            pyautogui.hotkey('shift','tab','enter')
            pyautogui.hotkey('alt', 'h')#line coloring
            pyautogui.press(['l','down','down','down','down','down','down','down','left','left','left','enter'])
            pyautogui.hotkey('alt', 'h')#fill coloring
            pyautogui.press(['i','down','down','down','down','down','down','down','left','left','left','enter'])
            #Options.font_white(self)
            AP_orange_count += 1   # adds one to total orange count
        print(f'Colored {AP_orange_count} orange APs')
        print("Select another color, or press 'QUIT' to end the program.\n")

    def auto_script_blue(self):   #finds AP name in Visio and colors it BLUE,
        Options.search_options(self)
        AP_blue_count = 0
        for x in AP_blue_list: 
            AP_num = AP_blue_list[AP_blue_count]
            AP_num_print = (AP_num[8:13]) if len(AP_num) > 5 else AP_num #Truncates site name, IDF, and AP letter.
            print(AP_num_print)
            pyautogui.hotkey('ctrl', 'f')
            pyautogui.typewrite(AP_num_print)
            pyautogui.press(['enter', 'esc','esc','tab','enter'])
            pyautogui.hotkey('shift','tab','enter')
            pyautogui.hotkey('shift','tab','enter')
            pyautogui.hotkey('alt', 'h')#line coloring
            pyautogui.press(['l','down','down','down','down','down','down','down','right','enter'])
            pyautogui.hotkey('alt', 'h')#fill coloring
            pyautogui.press(['i','down','down','down','down','down','down','down','right','enter'])
            #Options.font_white(self)
            AP_blue_count += 1      #adds one to total blue count
        print(f'Colored {AP_blue_count} blue APs')
        print("Select another color, or press 'QUIT' to end the program.\n")
    
    #def auto_script_total(self):
        #print(f'Total APs: ')
        #print(f'Total Failed APs:')

    def save_new_sheet(self):   #Pop-up message box, asks if user wants to save sorted date to new excell file/sheet
        result = messagebox.askyesno("Visio AP Coloring Tool","Do you want to save data in new Excell?")
        print(result)
        if result == True:
            print("data saved in 'New File'")
            
        else:
            pass


class AutoColor(tkinter.Frame):  #POP-UP GUI for choosing majority of Sheets color
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.pack()
        self.create_widgets()

    def create_widgets(self):   #buttons in the GUI
        self.title = tkinter.Label(self, font = 30, text = "Select a color to highlight:")
        self.title.pack(side = "top", pady = 30)

        self.press_grey = tkinter.Button(self, fg = "grey")    #Button, colored green
        self.press_grey["text"] = "GREY"                      #button, named green
        self.press_grey["command"] = self.run_grey              #when button pressed, execute output1
        self.press_grey.pack(side="top", pady = 10)                       #position selt at top most position

        self.press_green = tkinter.Button(self, text = "Green", fg = "green", command = self.run_green).pack(side="top")
        self.press_red = tkinter.Button(self, text = "Red", fg = "red", command = self.run_red).pack(side="top")
        self.press_orange = tkinter.Button(self, text = "Orange", fg = "orange", command = self.run_orange).pack(side="top")
        self.press_blue = tkinter.Button(self, text = "Blue", fg = "blue", command = self.run_blue).pack(side="top")
        self.quit = tkinter.Button(self, text="QUIT", command=self.close)
        self.quit.pack(side="bottom", pady = 30)

    def run_grey(self):   #colors all  green
        print("Setting all APs to Grey...")
        print("Please double-click into your Visio file.\n")
        time.sleep(3)
        Options.auto_script_grey(self)     
        #self.close()

    def run_green(self):
        print("Coloring all GREEN APs...")
        print("Please double-click into your Visio file.\n")
        time.sleep(3)
        #Options.font_white(self)  
        Options.auto_script_green(self)
        #self.close()

    def run_red(self):  #colors all  red
        print("Coloring all RED APs...")
        print("Please double-click into your Visio file.\n")
        time.sleep(3)
        #Options.font_white(self)  
        Options.auto_script_red(self)
        #self.close()

    def run_orange(self):  #colors all  orange
        print("Coloring all ORANGE APs...")
        print("Please double-click into your Visio file.\n")
        time.sleep(3)  
       # Options.font_white(self)   
        Options.auto_script_orange(self)     
        #self.close()

    def run_blue(self):  #colors all  blue
        print("Coloring all BLUE APs...")
        print("Please double-click into your Visio file.\n")
        time.sleep(3)
        #Options.font_white(self)  
        Options.auto_script_blue(self)     
        #self.close()
    
    def close(self):  #Ends the porgram when user selcts "QUIT"
        print("Closing program... Have a nice day   :)")
        time.sleep(1)
        sys.exit()


def main():
    #choice = "" #creates a blank string for user input
    keyPress = ""
    #while choice != "3": 
    while keyPress != "p": 
        print("Enter '[' to open Visio AP Naming tool")
        print("Enter ']' to open Visio AP Coloring tool")
        print("Enter 'p' to quit program\n")
        #choice = input(">>> ")
        keyPress = keyboard.read_key()
        print(f'>>> {keyPress}')
        if keyPress == "[":
            pyautogui.press('backspace')
            print("Starting Visio Naming Tool...")           
            theGUI()
            #vnt.theGUI()    #the GUI function calls these internally:
            break               #vnt.startWorkbook()   
                                    #vnt.openFiles()
                                    #vnt.visioLoop()
                                        #vnt.visioGuts()                                      
        elif keyPress == "]":           #vnt.saveExcel()   
            #AutoColor Tool
            pyautogui.press('backspace')
            print("Starting Visio Color Tool...")
            start = Options()
            start.open_audit_sheet()
            start.audit_sheet_sorter()
            root = tkinter.Tk()
            A_C = AutoColor(master=root)
            A_C.master.title("Visio AP Coloring Tool")
            A_C.mainloop()
            return
        
        elif keyPress == "p":
            pyautogui.press('backspace')
            print("Closing program... Have a nice day   :)")
            break

        else:
            print("Invalid input")
            pass


if __name__ == "__main__":
    main()   #this is the call to run the whole program


#added functionality for option #8 - red, blue, and orange
#renamed varibales in AutoColor Class
#added fucntionality to color APs green
#added functionality to color ALL APs grey at the start, must be used for each sheet
#Fixed exit errors, used "sys.exit()"
#Note, need to edit the varible repetitiveness of Count, count list, list 
#fixed error that was caused from running the porgram after closing it