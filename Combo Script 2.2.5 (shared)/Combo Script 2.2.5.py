import keyboard, openpyxl, pyautogui, tkinter, time, sys, os, datetime
from tkinter import filedialog, messagebox, Tk
from openpyxl import Workbook, load_workbook
from guizero import App, Text, TextBox, PushButton, Window

##### A Soldat and Luxo Program #####

global siteName
global apPrefix
global apNumber
global apName
global excelCount
global excelNumber
global rowLetter
global rowNumber
global rowWholeName
global list
global excelName
global formatNumber
global excelStartingInteger
global keyPress

siteName = str ('Paul')
apPrefix = siteName + '-XX-'
apNumber =  int('1')
apName = str(apPrefix) + str(apNumber)
excelCount = int('1')
excelNumber = int('1')
rowLetter = 'A'
rowNumber = int('1')

rowWholeName = str(rowLetter) + str(rowNumber)
list=[99999]
excelName = ''
formatNumber = format(apNumber, '05')
excelStartingInteger = int('1')
currentTime = datetime.datetime.now().strftime("%Y-%m-%d  %H.%M.%S")


def startWorkbook ():#Starts an excel. Required for visioTool.
    global wb
    wb = Workbook()
    global ws
    ws = wb.active

def openFiles():#Opens file explorer
    global excelName
    #Tk().withdraw() # We don't want a full GUI, so keep the root window from appearing
    excelName = filedialog.askopenfilename() # Show an "Open" dialog box and return the path to the selected file
    print(f'Selected save file: {excelName}\n')
    print("===========================================================================\n")

def pause(): #Pause logic
    keyPress = keyboard.read_key()
    if keyPress == 'pause': #Reads if pause button has been pressed and unpauses
        print('Unpaused')
        time.sleep(.3)
        exit
    else: #If pause has not been pressed then loop repeats infinitely WOO
        pause()
        
def theGUI(): #all the GUI stuff
    print("INSTRUCTIONS:")
    #print("  > Set Up - use the Visio Naming Tool to change:")
   # print("      > SITE name (use four capital letters)")
    #print("          > 'Confirm'")
   # print("      > AP Starting Number (numerical values only)")
    #print("          > 'Confirm'")
    #print("      > Push 'GO'")
   # print("          > Select an excel file to save the AP names\n")
    print("  > Functionality:")
    print("      > Each of the following hotkeys will add 1 AP to the excel file")
    print("          > Press ' ` ' to add a default AP")
    print("          > Press ' d ' to add a door AP")
    print("          > Press ' e ' to add a external AP")
    print("          > Press ' g ' to add a guard shack AP")
    print("          > Press ' h ' to add a high racking AP")
    print("          > Press ' m ' to add a MOD/KIVA AP")
    print("          > Press ' s ' to add a standup AP")
    print("      > If you make an error and need to make a fix:")
    print("          > Press ' [ ' to navigate down 1 AP")
    print("          > Press ' ] ' to navigate up 1 AP")
    print("      > To save:")
    print("          > Press ' = ' to save AP IDs into chosen excel")
    #print("      > To quit:")
    #print("          > Press ' q ' to quit the program\n")
    print("===========================================================================\n")

    
    def apStartingNumber(): # part of GUI that allows changing of the starting AP number
        global apNumber
        global formatNumber
        apNumber = int(startingNumber.value)
        formatNumber = format(apNumber, '05')
        print ('AP number changed to: ' + str(apNumber))
        changingText.value = "AP number changed to: " + str(apNumber) #Text for changing AP number

    def changeSiteName(): # part of GUI that allows changing of the site name
        global siteName
        global apPrefix
        global apName
        siteName = str(siteNames.value)
        apPrefix = siteName + '-XX-'
        apName = str(apPrefix) + str(apNumber)
        print('Site name changed to: ' + siteName)
        changingText.value = "Site changed to: " + siteName #Text for changing site name
        
    def directions(): #Once "Go?" has been pressed execute the following items
        startWorkbook()
        openFiles()
        app.hide()
        visioLoop()
        python = sys.executable #Restarts the whole program
        os.execl(python, python, * sys.argv) #Restarts the whole program
### VNT GUI ###      
    app = App(title = "Phoenix Oath", width=352, height=132, layout='grid')

    button7 = PushButton(app, text = "Go?", command = directions, grid=[2,3])

    #Logic for changing starting AP number
    startingNumberText = Text(app, text="AP Number?", align="left", grid=[0,1]) #Text asks for AP number
    startingNumber = TextBox(app, align="right",text = "1", width=30, grid=[1,1]) #Text box for data entry
    button4 = PushButton(app, text = "Confirm", command = apStartingNumber, grid=[2,1])

    #Logic for changing site name
    siteNamesText = Text(app, text="Site Name?", align="left",  grid=[0,2]) #Text asks for site name
    siteNames = TextBox(app, align="right",text = "SITE", width=30, grid=[1,2]) #Text box for data entry
    button6 = PushButton(app, text = "Confirm", command = changeSiteName, grid=[2,2])

    changingText = Text(app,text="War has changed", align ="left", grid=[1,3])
    app.display() # initiates the GUI. Allowing it to be used
    
def saveExcel(): # saves Visio Tool names to Excel
    global displayText
    global excelName
    wb.save(excelName) # Saves workbook
    print ('\nWorksheet saved')
    #print ('End of Visio Tool')
        
def visioGuts(): # the internals to the Visio Tool. Determines how most of the program is run
    global apNumber
    global formatNumber
    global excelStartingInteger
    print (f'     {apName}')
    ws.cell(excelStartingInteger, 1, apName)  #writes in excel **format** ->(row, column, content to be written in cell)
    pyautogui.press('backspace'); pyautogui.typewrite(str(formatNumber)) # takes control of keyboard. hits backspace and types AP number
    apNumber += 1 #increments ap number up by 1
    excelStartingInteger += 1 # Increments Excel cell to be written in
    formatNumber = format(apNumber, '05') # Modifies apNumber by adding up to 5 zeros in front

def visioLoop():
    global apNumber
    global formatNumber
    global excelStartingInteger
    global apName

    
    print("Access points named: ")

    while True:
        keyPress = keyboard.read_key() #reads ALL keypress's and saves to variable
        
        if keyPress == '`' : #Adds a Default Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber)
            visioGuts ()
            
        elif keyPress == 's': #Adds a Standup Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber) + 'S'
            visioGuts ()
            
        elif keyPress == 'g' : #Adds a Guard Shack Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber) + 'G'
            visioGuts ()
            
        elif keyPress == 'm' : #Adds a MOD/KIVA Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber) + 'M'
            visioGuts ()
            
        elif keyPress == 'h' : #Adds a High Racking Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber) + 'H'
            visioGuts ()
            
        elif keyPress == 'd' : #Adds a Door Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber) + 'D'
            visioGuts ()

        elif keyPress == 'e' : #Adds a External Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber) + 'E'
            visioGuts ()
            
        elif keyPress == '[': #goes down one ap number.
            apNumber -= 1
            excelStartingInteger -= 1
            formatNumber = format(apNumber, '05')
            apName = str(apPrefix) + str(formatNumber)
            print (f'Fix: {apName} ?')
            time.sleep(.09)

        elif keyPress == ']': #goes up one ap number.
            apNumber += 1
            excelStartingInteger += 1
            formatNumber = format(apNumber, '05')
            apName = str(apPrefix) + str(formatNumber)
            print (f'Fix: {apName} ?')
            time.sleep(.09)
            
        elif keyPress == 'pause': #Well it pauses everthing.....
            print('Paused')
            time.sleep(.3)
            pause() #Program is stuck in the pause loop until pause is pressed again

        elif keyPress == '=': #Saves the CAD Cutsheet   
            saveExcel()
            
        elif keyPress == 'q':  #closes the Visio tool
            #sys.exit()
            exit()


class Options:
    def open_audit_sheet (self): #Opens a file explorer and returns path from chosen file
        global audit_sheet
        print("Select the specific 'Audit Sheets' to use as reference for coloring")
        Tk().withdraw() # We don't want a full GUI, so keep the root window from appearing
        xlname = filedialog.askopenfilename() # Show an "Open" dialog box and return the path to the selected file
        print(f'Loading excel file: {xlname}')
        audit_book = load_workbook(xlname)#Opens Excel
        audit_sheet = audit_book['Audit'] #grabs information from the "audit" sheet

    def audit_sheet_sorter (self): #reads all AP's from chosen Audit Sheet and categories them by their color
        global site
        global name
        global AP_red_list
        global AP_blue_list
        global AP_green_list
        global AP_orange_list
        global AP_grey_list
        
        xlname_col = 'A' #column letter to read AP names from
        xlname_col_fail = 'K' #column letter to read fails from
        xlname_row = int('3') #Starting row in excel
        xlname_row_fail = int('3') #Starting row in excel
    
        site = []  #makes a blank array to hold the SITE string value
        AP_red_list = []  #initiates a list for green AP's ('2')
        AP_blue_list = []  #initiates a list for green AP's ('4')
        AP_green_list = []  #initiates a list for green AP's ('1')
        AP_orange_list = []  #initiates a list for green AP's ('3')
        AP_grey_list = []
        
        AP_red_count_list = int('0')#Keeps track how long the list is
        AP_blue_count_list = int('0')
        AP_green_count_list = int('0')
        AP_orange_count_list = int('0') 
        AP_total_count_list = int('0')     
        AP_grey_count_list = int('0')     
                
        for AP in range (0, audit_sheet.max_row +1): #range of APS being colored
            xlname_col_fail_num = str(xlname_col_fail) + str(xlname_row_fail)
            xlname_col_num = str(xlname_col) + str(xlname_row)
            cell_value1 = audit_sheet[xlname_col_fail_num]
            cell_value2 = audit_sheet[xlname_col_num]
            AP_fail = cell_value1.value
            AP_ID = cell_value2.value

            site.append(AP_ID)  #appending cell values of first column to Site
            site_name = site[0] #opens that the first cell in the column
            name = site_name[0:4] #read the first 4 characters in cell value (site name)
        
            if AP_fail == 1: #reads all green AP's
                AP_green_list.append(AP_ID)
                #print(AP_green_list[AP_green_count_list])
                AP_green_count_list += 1
                AP_total_count_list += 1

            elif AP_fail == 2: #reads all red AP's
                AP_red_list.append(AP_ID)
                #print(AP_red_list[AP_red_count_list])
                AP_red_count_list += 1
                AP_total_count_list += 1
                AP_grey_list.append(AP_ID)
                AP_grey_count_list += 1
                
            elif AP_fail == 3: #reads all orange AP's
                AP_orange_list.append(AP_ID)
                #print(AP_orange_list[AP_orange_count_list])
                AP_orange_count_list += 1
                AP_total_count_list += 1
                AP_grey_list.append(AP_ID)
                AP_grey_count_list += 1
                
            elif AP_fail == 4: #reads all blue AP's
                AP_blue_list.append(AP_ID)
                #print(AP_blue_list[AP_blue_count_list])
                AP_blue_count_list += 1
                AP_total_count_list += 1
                AP_grey_list.append(AP_ID)
                AP_grey_count_list += 1

            elif AP_fail == 5: #reads all red and orange AP's
                AP_red_list.append(AP_ID)
                #print(AP_red_list[AP_red_count_list])
                AP_red_count_list += 1
                AP_orange_list.append(AP_ID)
                #print(AP_orange_list[AP_orange_count_list])
                AP_orange_count_list += 1
                AP_total_count_list += 1
                AP_grey_list.append(AP_ID)
                AP_grey_count_list += 1

            elif AP_fail == 6: #reads all red and blue AP's
                AP_red_list.append(AP_ID)
                #print(AP_red_list[AP_red_count_list])
                AP_red_count_list += 1
                AP_blue_list.append(AP_ID)
                #print(AP_blue_list[AP_blue_count_list])
                AP_blue_count_list += 1
                AP_total_count_list += 1
                AP_grey_list.append(AP_ID)
                AP_grey_count_list += 1

            elif AP_fail == 7: #reads all orange and blue AP's
                AP_orange_list.append(AP_ID)
                #print(AP_orange_list[AP_orange_count_list])
                AP_orange_count_list += 1
                AP_blue_list.append(AP_ID)
                #print(AP_blue_list[AP_blue_count_list])
                AP_blue_count_list += 1
                AP_total_count_list += 1
                AP_grey_list.append(AP_ID)
                AP_grey_count_list += 1

            elif AP_fail == 8: #reads all red and orange AP's
                AP_red_list.append(AP_ID)
                #print(AP_red_list[AP_red_count_list])
                AP_red_count_list += 1
                AP_orange_list.append(AP_ID)
                #print(AP_orange_list[AP_orange_count_list])
                AP_orange_count_list += 1    
                AP_blue_list.append(AP_ID)
                #print(AP_blue_list[AP_blue_count_list])
                AP_blue_count_list += 1
                AP_total_count_list += 1
                AP_grey_list.append(AP_ID)
                AP_grey_count_list += 1

            else:   #at end of Audit Sheet, program exits loop, prints results
                break  
            xlname_row_fail += 1
            xlname_row += 1
        
        print("\n==============================================================================\n")
        print(f'{name} Audit Sheet sorted\n')   #stats, used to help user understand the audit
        print(f'    Surveyed APs:       {AP_total_count_list}')
        print(f'    Passed (Green) APs: {AP_green_count_list} ')
        print(f'    Failed (Grey) APs:  {AP_grey_count_list}')
        print(f'     - Red APs:         {AP_red_count_list} ')
        print(f'     - Blue APs:        {AP_blue_count_list} ')
        print(f'     - Orange APs:      {AP_orange_count_list}\n')
        
        Options.instructions_starting(self)  #print the main instructions of the program

    def search_options (self):  #sets visio to search all pages for AP IDs, must be done before coloring
        #print (">>> Starting Search...\n")
        print("\nColoring, please standby... \n")
        pyautogui.hotkey('ctrl', 'f')
        pyautogui.press(['tab','tab'])
        pyautogui.press('down')
        pyautogui.press('esc')
        
    def auto_script_grey(self):   
        AP_grey_count = 0
        while True:
            if keyboard.is_pressed('enter'):        
                Options.search_options(self)
                for x in AP_grey_list: 
                    AP_num = AP_grey_list[AP_grey_count]
                    AP_num_print = (AP_num[8:13]) if len(AP_num) > 5 else AP_num #Truncates site name, IDF, and AP letter.
                    #print(f'     {AP_num_print}')
                    pyautogui.hotkey('ctrl', 'f')
                    pyautogui.typewrite(AP_num_print)
                    pyautogui.press(['enter', 'esc','esc','tab','enter'])
                    pyautogui.hotkey('shift','tab','enter')
                    pyautogui.hotkey('shift','tab','enter')
                    pyautogui.hotkey('alt', 'h')#line coloring
                    pyautogui.press(['l','down','down','down','down','down','left','left','left','enter'])
                    pyautogui.hotkey('alt', 'h')#fill coloring
                    pyautogui.press(['i','down','down','down','down','down','left','left','left','enter'])
                    AP_grey_count += 1     #adds one to total grey count
                    if keyboard.is_pressed('esc'):   #escape sequence if user wants to quit
                        print("Stopping color sequence")
                        break
                    else:   #if esc is not pressed, the for loop will continue iterating
                        pass
                pyautogui.press('esc')
                break
            if keyboard.is_pressed('esc'):
                break
        print(f'\nFinished coloring {AP_grey_count} grey APs')
        print("\n==============================================================================\n")
        #Options.instructions_saving(self)

    def auto_script_green(self):  #finds AP name in Visio and colors it GREEN,
        AP_green_count = 0
        while True:
            if keyboard.is_pressed('enter'):
                Options.search_options(self)  
                for x in AP_green_list: 
                    AP_num = AP_green_list[AP_green_count]
                    AP_num_print = (AP_num[8:13]) if len(AP_num) > 5 else AP_num #Truncates site name, IDF, and AP letter.
                    #print(f'     {AP_num_print}')
                    pyautogui.hotkey('ctrl', 'f')
                    pyautogui.typewrite(AP_num_print)
                    pyautogui.press(['enter', 'esc','esc','tab','enter'])
                    pyautogui.hotkey('shift','tab','enter')
                    pyautogui.hotkey('shift','tab','enter')
                    pyautogui.hotkey('alt', 'h')#line coloring
                    pyautogui.press(['l','down','down','down','down','down','down','down','left','enter'])
                    pyautogui.hotkey('alt', 'h')#fill coloring
                    pyautogui.press(['i','down','down','down','down','down','down','down','left','enter'])
                    AP_green_count += 1     #adds one to total grey count
                    if keyboard.is_pressed('esc'):
                        print("Stopping color sequence")
                        break
                    else:
                        pass
                pyautogui.press('esc')
                break
            if keyboard.is_pressed('esc'):
                break
        print(f'\nFinished coloring {AP_green_count} green APs')
        print("\n==============================================================================\n")
        #Options.instructions_saving(self)

    def auto_script_red(self): #finds AP name in Visio and colors it RED,   
        AP_red_count = 0 
        while True:
            if keyboard.is_pressed('enter'):
                Options.search_options(self)   
                for x in AP_red_list: #finds AP name in Visio and colors it
                    AP_num = AP_red_list[AP_red_count]
                    AP_num_print = (AP_num[8:13]) if len(AP_num) > 5 else AP_num #Truncates site name, IDF, and AP letter.
                    #print(f'     {AP_num_print}')
                    pyautogui.hotkey('ctrl', 'f')
                    pyautogui.typewrite(AP_num_print)
                    pyautogui.press(['enter', 'esc','esc','tab','enter'])
                    pyautogui.hotkey('shift','tab','enter')
                    pyautogui.hotkey('shift','tab','enter')
                    pyautogui.hotkey('alt', 'h')#line coloring
                    pyautogui.press(['l','down','down','down','down','down','down','down','left','left','left','left','enter'])
                    pyautogui.hotkey('alt', 'h')#fill coloring
                    pyautogui.press(['i','down','down','down','down','down','down','down','left','left','left','left','enter'])
                    AP_red_count += 1   #adds one to total red count
                    if keyboard.is_pressed('esc'):
                        print("Stopping color sequence")
                        break
                    else:
                        pass
                pyautogui.press('esc')
                break
            if keyboard.is_pressed('esc'):
                break
        print(f'\nFinished coloring {AP_red_count} red APs')

    def auto_script_orange(self): #finds AP name in Visio and colors it ORANGE, 
        AP_orange_count = 0
        while True:
            if keyboard.is_pressed('enter'):
                Options.search_options(self)
                for x in AP_orange_list: #finds AP name in Visio and colors it
                    AP_num = AP_orange_list[AP_orange_count]
                    AP_num_print = (AP_num[8:13]) if len(AP_num) > 5 else AP_num #Truncates site name, IDF, and AP letter.
                    #print(f'     {AP_num_print}')
                    pyautogui.hotkey('ctrl', 'f')
                    pyautogui.typewrite(AP_num_print)
                    pyautogui.press(['enter', 'esc','esc','tab','enter'])
                    pyautogui.hotkey('shift','tab','enter')
                    pyautogui.hotkey('shift','tab','enter')
                    pyautogui.hotkey('alt', 'h')#line coloring
                    pyautogui.press(['l','down','down','down','down','down','down','down','left','left','left','enter'])
                    pyautogui.hotkey('alt', 'h')#fill coloring
                    pyautogui.press(['i','down','down','down','down','down','down','down','left','left','left','enter'])
                    AP_orange_count += 1   # adds one to total orange count
                    if keyboard.is_pressed('esc'):
                        print("Stopping color sequence")
                        break
                    else:
                        pass
                pyautogui.press('esc')
                break
            if keyboard.is_pressed('esc'):
                break
        print(f'\nFinished coloring {AP_orange_count} orange APs')

    def auto_script_blue(self):   #finds AP name in Visio and colors it BLUE,
        AP_blue_count = 0
        while True:
            if keyboard.is_pressed('enter'):
                Options.search_options(self)
                for x in AP_blue_list: 
                    AP_num = AP_blue_list[AP_blue_count]
                    AP_num_print = (AP_num[8:13]) if len(AP_num) > 5 else AP_num #Truncates site name, IDF, and AP letter.
                    #print(f'     {AP_num_print}')
                    pyautogui.hotkey('ctrl', 'f')
                    pyautogui.typewrite(AP_num_print)
                    pyautogui.press(['enter', 'esc','esc','tab','enter'])
                    pyautogui.hotkey('shift','tab','enter')
                    pyautogui.hotkey('shift','tab','enter')
                    pyautogui.hotkey('alt', 'h')#line coloring
                    pyautogui.press(['l','down','down','down','down','down','down','down','right','enter'])
                    pyautogui.hotkey('alt', 'h')#fill coloring
                    pyautogui.press(['i','down','down','down','down','down','down','down','right','enter'])
                    AP_blue_count += 1      #adds one to total blue count
                    if keyboard.is_pressed('esc'):
                        print("Stopping color sequence")
                        break
                    #if keyboard.is_pressed('pause'):
                    #   pause()
                    else:
                        pass
                pyautogui.press('esc')
                break
            if keyboard.is_pressed('esc'):
                break
           # if keyboard.is_pressed('esc'):
           #     print("Stopped!")
           #     break
        print(f'\nFinished coloring {AP_blue_count} blue APs')
          
    def instructions_starting(self):
        print("INSTRUCTIONS - set up")
        print(" > Duplicate your original Visio Design Sheets, rename: 'Red'")
        print(" > Set all AP markers' text font to WHITE ")
        print(" > Set all border and fill line color to either GREY or GREEN,")
        print("        (depending on the majority, reference above stats)")
        #print("     > Repeat this step for each floor")
        print("\n > Using the pop-up 'Visio AP Illuminator' tool, select either GREEN or GREY,")
        print("        (whichever one you did NOT use for the majority)")
        print("     > Follow the on-screen steps from there\n")
        #print(" > Exit that Visio and duplicate it twice,")
        #print("     > Rename the copies: 'Orange' and 'Blue'\n")
        #print(" > Open any of the new Visio files:")
        #print("     > Use the 'Visio AP Illuminator' tool")
        #print("         > Select the respective color you want to use")
        #print("     > Follow the on-screen steps from there")
        #print("     > Repeat for each Visio and their respective color\n")
        print("==============================================================================\n")
        
    def instructions_copying(self):
        print("\nINSTRUCTIONS - copying")
        print(" > Save and Exit this Visio then duplicate it twice,")
        print("     > Rename the copies: 'Orange' and 'Blue'\n")
        print(" > Open any of the new Visio files:")
        print("     > Use the 'Visio AP Illuminator' tool")
        print("         > Select the respective color you want to use")
        print("     > Follow the on-screen steps from there")
        print("     > Repeat for each Visio and their respective color")
         
    def instructions_saving(self):
        #print(" > To Save as an image for Report Generation, go to:")
        #print("     >  File --> Export --> Change File Type --> SVG (Scalable Vector Graphic)")
        #print("\nINSTRUCTIONS - saving")
        print("\n > Press ' = ' to Save as an SVG image for Report Generation, ")
        print("         > Destination: same as Visio")
        print("         > File Name: 'Color' - 'scope'   Ex. 'Red - A1'")
       # print("     > Repeat for each floor")
        print("\n > Press ' esc ' to return to the 'Visio AP Illuminator' tool")
        print(" > Open another Visio file for a different color or press 'QUIT'\n")
        print("==============================================================================\n")
        while True:    
            if keyboard.is_pressed('='):
                #pyautogui.press(['backspace', 'esc'])
                pyautogui.hotkey('alt', 'f')
                pyautogui.press(['e', 'c','g','enter'])
            elif keyboard.is_pressed('esc'):
                break
        #print("Select another color or press 'QUIT'\n")

    def instructions_coloring(self):
        print(" > Click into your Visio file")
        print(" > Press ' enter ' to start the color sequence")
        print("     > While the program is running,")
        print("           DO NOT click anywhere outide of the Visio file")
        print(" > If you chose the wrong color, press ' esc ' to return to 'AP Illuminator'\n")
        print(" >  Spam ' esc ' while the program is running if you need to exit early  ")
        

    ####    #def save_new_excel(self):   #Pop-up message box, asks if user wants to save sorted date to new excel file/sheet
        #   result = messagebox.askyesno("Visio AP Coloring Tool","Do you want to save data in new excel?")
        #    print(result)
        #    if result == True:
        #        print("data saved in 'New File'")  
        #    else:
        #        pass


class AutoColor(tkinter.Frame):  #POP-UP GUI for choosing majority of Sheets color
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.grid(column = 0, row = 0)
        self.create_widgets()

    def create_widgets(self):   #buttons in the GUI
        self.title = tkinter.Label(self, font = 30, text = "Select a color to highlight:").grid(column=2, row =1, pady = 30, ipadx = 30)

       # self.entry_grey = tkinter.Label(self, text = "Page Count: ").grid(column=1, row =2)
       # self.entry_grey = tkinter.Entry(self, width = 5).grid(column=2, row =2)
        self.press_grey = tkinter.Button(self, text = "Grey", fg = "grey", command = self.run_grey).grid(column=2, row =3)     
        self.press_green = tkinter.Button(self, text = "Green", fg = "green", command = self.run_green).grid(column=2, row =4, pady = 5) 
        self.press_red = tkinter.Button(self, text = "Red", fg = "red", command = self.run_red).grid(column=2, row =5) 
        self.press_orange = tkinter.Button(self, text = "Orange", fg = "orange", command = self.run_orange).grid(column=2, row =6, pady = 5)
        self.press_blue = tkinter.Button(self, text = "Blue", fg = "blue", command = self.run_blue).grid(column=2, row =7)
        self.quit = tkinter.Button(self, text="QUIT", command=self.close).grid(column=2, row =8, pady = 30) 
    
    def run_grey(self):   #colors all  green
        print("Color selected: GREY\n")
        Options.instructions_coloring(self)
        Options.auto_script_grey(self) 
        #Options.instructions_copying(self)    
        #Options.instructions_saving(self)

    def run_green(self):
        print("Color selected: GREEN\n")
        Options.instructions_coloring(self)
        Options.auto_script_green(self)
        #Options.instructions_copying(self)
        #Options.instructions_saving(self)

    def run_red(self):  #colors all  red
        print("Color selected: RED\n")
        Options.instructions_coloring(self)
        Options.auto_script_red(self)
        Options.instructions_saving(self)

    def run_orange(self):  #colors all  orange
        print("Color selected: ORANGE\n")
        Options.instructions_coloring(self)
        Options.auto_script_orange(self) 
        Options.instructions_saving(self)    

    def run_blue(self):  #colors all  blue
        print("Color selected: BLUE\n")
        Options.instructions_coloring(self)
        Options.auto_script_blue(self)     
        Options.instructions_saving(self)

    def close(self):  #Ends the porgram when user selcts "QUIT"
        result = messagebox.askyesno("Visio AP Coloring Tool","Are you sure?")
        if result == True:
            #print("Restarting Program...")
            #start = Options()
            #start.open_audit_sheet()
            #start.audit_sheet_sorter()
            print("Closing program... Have a nice day   :)")
            time.sleep(2)
            #sys.exit()  #for exceutable quitting
            sys.exit()  #for VSC coding
        else:
            #print("Closing program... Have a nice day   :)")
            #time.sleep(1)
            #sys.exit()
            pass
        
        
def main():
    keyPress = "" 
    print("Enter '[' to open Visio AP Naming tool")
    print("Enter ']' to open Visio AP Coloring tool")
    #print("Enter 'p' to quit program\n")
    while keyPress != "p": 
        #choice = input(">>> ")
        keyPress = keyboard.read_key()
        #print(f'>>> {keyPress}')
        if keyPress == "[":
            pyautogui.press('backspace')
            print("Starting Visio Naming Tool... \n")           
            theGUI()
            #vnt.theGUI()    #the GUI function calls these internally:
            break               #vnt.startWorkbook()   
                                    #vnt.openFiles()
                                    #vnt.visioLoop()
                                        #vnt.visioGuts()                                      
        elif keyPress == "]":           #vnt.saveExcel()   
            #AutoColor Tool
            pyautogui.press('backspace')
            print("Starting Visio Color Tool... \n")
            start = Options()
            start.open_audit_sheet()
            start.audit_sheet_sorter()
            root = tkinter.Tk()
            AC = AutoColor(master=root)
            AC.master.title('AP Illuminator')
            AC.mainloop()
            return
        elif keyPress == "p":
            pyautogui.press('backspace')
            print("Closing program... Have a nice day   :)")
            break
        else:
            print("Invalid input")
            pass


if __name__ == "__main__":
    main()   #this is the call to run the whole program



# ******************************
# Changes *********************
# ******************************
# Removed color ALL grey function for color individual greys
# Added an escape feature to leave the visio if user wants to do so early
# Added in depth instructions for both Visio Naming and Coloring
# fixed wording n Visio name tool
# Added hotkey fucntion for saving visio to SVG, scalable vector graphic
# Coloring now operates when user presses 'enter'
# Commented out a lot of code for final publishing of this version
# 
# 
# 
