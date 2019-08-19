import keyboard, openpyxl, pyautogui, tkinter, time, sys, os, datetime
from tkinter import filedialog, messagebox, Tk, ttk
from openpyxl import Workbook, load_workbook
from guizero import App, Text, TextBox, PushButton, Window

##### A Soldat and Luxo Program #####

global siteName
global apPrefix
global apNumber
global apName
global excelCount
global excelNumber
global rowLetter
global rowNumber
global rowWholeName
global list
global excelName
global formatNumber
global excelStartingInteger
global keyPress

siteName = str ('Paul')
apPrefix = siteName + '-XX-'
apNumber =  int('1')
apName = str(apPrefix) + str(apNumber)
excelCount = int('1')
excelNumber = int('1')
rowLetter = 'A'
rowNumber = int('1')

rowWholeName = str(rowLetter) + str(rowNumber)
list=[99999]
excelName = ''
formatNumber = format(apNumber, '05')
excelStartingInteger = int('1')
currentTime = datetime.datetime.now().strftime("%Y-%m-%d  %H.%M.%S")


def startWorkbook ():#Starts an excel. Required for visioTool.
    global wb
    wb = Workbook()
    global ws
    ws = wb.active

def openFiles():#Opens file explorer
    global excelName
    #Tk().withdraw() # We don't want a full GUI, so keep the root window from appearing
    excelName = filedialog.askopenfilename() # Show an "Open" dialog box and return the path to the selected file
    print(f'Selected save file: {excelName}\n')
    print("===========================================================================\n")

def pause(): #Pause logic
    keyPress = keyboard.read_key()
    if keyPress == 'pause': #Reads if pause button has been pressed and unpauses
        print('Unpaused')
        time.sleep(.3)
        exit
    else: #If pause has not been pressed then loop repeats infinitely WOO
        pause()
        
def theGUI(): #all the GUI stuff
    print("INSTRUCTIONS:")
    #print("  > Set Up - use the Visio Naming Tool to change:")
   # print("      > SITE name (use four capital letters)")
    #print("          > 'Confirm'")
   # print("      > AP Starting Number (numerical values only)")
    #print("          > 'Confirm'")
    #print("      > Push 'GO'")
   # print("          > Select an excel file to save the AP names\n")
    print("  > Functionality:")
    print("      > Each of the following hotkeys will add 1 AP to the excel file")
    print("          > Press ' ` ' to add a default AP")
    print("          > Press ' d ' to add a door AP")
    print("          > Press ' e ' to add a external AP")
    print("          > Press ' g ' to add a guard shack AP")
    print("          > Press ' h ' to add a high racking AP")
    print("          > Press ' m ' to add a MOD/KIVA AP")
    print("          > Press ' s ' to add a standup AP")
    print("      > If you make an error and need to make a fix:")
    print("          > Press ' [ ' to navigate down 1 AP")
    print("          > Press ' ] ' to navigate up 1 AP")
    print("      > To save:")
    print("          > Press ' = ' to save AP IDs into chosen excel")
    #print("      > To quit:")
    #print("          > Press ' q ' to quit the program\n")
    print("===========================================================================\n")

    
    def apStartingNumber(): # part of GUI that allows changing of the starting AP number
        global apNumber
        global formatNumber
        apNumber = int(startingNumber.value)
        formatNumber = format(apNumber, '05')
        print ('AP number changed to: ' + str(apNumber))
        changingText.value = "AP number changed to: " + str(apNumber) #Text for changing AP number

    def changeSiteName(): # part of GUI that allows changing of the site name
        global siteName
        global apPrefix
        global apName
        siteName = str(siteNames.value)
        apPrefix = siteName + '-XX-'
        apName = str(apPrefix) + str(apNumber)
        print('Site name changed to: ' + siteName)
        changingText.value = "Site changed to: " + siteName #Text for changing site name
        
    def directions(): #Once "Go?" has been pressed execute the following items
        startWorkbook()
        openFiles()
        app.hide()
        visioLoop()
        python = sys.executable #Restarts the whole program
        os.execl(python, python, * sys.argv) #Restarts the whole program
    ### VNT GUI ###      
    app = App(title = "Phoenix Oath", width=352, height=132, layout='grid')

    button7 = PushButton(app, text = "Go?", command = directions, grid=[2,3])

    #Logic for changing starting AP number
    startingNumberText = Text(app, text="AP Number?", align="left", grid=[0,1]) #Text asks for AP number
    startingNumber = TextBox(app, align="right",text = "1", width=30, grid=[1,1]) #Text box for data entry
    button4 = PushButton(app, text = "Confirm", command = apStartingNumber, grid=[2,1])

    #Logic for changing site name
    siteNamesText = Text(app, text="Site Name?", align="left",  grid=[0,2]) #Text asks for site name
    siteNames = TextBox(app, align="right",text = "SITE", width=30, grid=[1,2]) #Text box for data entry
    button6 = PushButton(app, text = "Confirm", command = changeSiteName, grid=[2,2])

    changingText = Text(app,text="War has changed", align ="left", grid=[1,3])
    app.display() # initiates the GUI. Allowing it to be used
    
def saveExcel(): # saves Visio Tool names to Excel
    global displayText
    global excelName
    wb.save(excelName) # Saves workbook
    print ('\nWorksheet saved')
    #print ('End of Visio Tool')
        
def visioGuts(): # the internals to the Visio Tool. Determines how most of the program is run
    global apNumber
    global formatNumber
    global excelStartingInteger
    print (f'     {apName}')
    ws.cell(excelStartingInteger, 1, apName)  #writes in excel **format** ->(row, column, content to be written in cell)
    pyautogui.press('backspace'); pyautogui.typewrite(str(formatNumber)) # takes control of keyboard. hits backspace and types AP number
    apNumber += 1 #increments ap number up by 1
    excelStartingInteger += 1 # Increments Excel cell to be written in
    formatNumber = format(apNumber, '05') # Modifies apNumber by adding up to 5 zeros in front

def visioLoop():
    global apNumber
    global formatNumber
    global excelStartingInteger
    global apName

    
    print("Access points named: ")

    while True:
        keyPress = keyboard.read_key() #reads ALL keypress's and saves to variable
        
        if keyPress == '`' : #Adds a Default Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber)
            visioGuts ()
            
        elif keyPress == 's': #Adds a Standup Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber) + 'S'
            visioGuts ()
            
        elif keyPress == 'g' : #Adds a Guard Shack Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber) + 'G'
            visioGuts ()
            
        elif keyPress == 'm' : #Adds a MOD/KIVA Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber) + 'M'
            visioGuts ()
            
        elif keyPress == 'h' : #Adds a High Racking Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber) + 'H'
            visioGuts ()
            
        elif keyPress == 'd' : #Adds a Door Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber) + 'D'
            visioGuts ()

        elif keyPress == 'e' : #Adds a External Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber) + 'E'
            visioGuts ()
            
        elif keyPress == '[': #goes down one ap number.
            apNumber -= 1
            excelStartingInteger -= 1
            formatNumber = format(apNumber, '05')
            apName = str(apPrefix) + str(formatNumber)
            print (f'Fix: {apName} ?')
            time.sleep(.09)

        elif keyPress == ']': #goes up one ap number.
            apNumber += 1
            excelStartingInteger += 1
            formatNumber = format(apNumber, '05')
            apName = str(apPrefix) + str(formatNumber)
            print (f'Fix: {apName} ?')
            time.sleep(.09)
            
        elif keyPress == 'pause': #Well it pauses everthing.....
            print('Paused')
            time.sleep(.3)
            pause() #Program is stuck in the pause loop until pause is pressed again

        elif keyPress == '=': #Saves the CAD Cutsheet   
            saveExcel()
            
        elif keyPress == 'q':  #closes the Visio tool
            #sys.exit()
            exit()


class Options:
    def open_audit_sheet (self): #Opens a file explorer and returns path from chosen file
        global audit_sheet
        global guts_sheet
        global audit_book     

        print("Select the specific 'Audit Sheets' to use as reference for coloring")
        Tk().withdraw() # We don't want a full GUI, so keep the root window from appearing
        xlname = filedialog.askopenfilename() # Show an "Open" dialog box and return the path to the selected file
        print(f'Loading excel file: {xlname}')
        audit_book = load_workbook(xlname)#Opens Excel
        audit_sheet = audit_book['Audit'] #grabs information from the "audit" sheet
        guts_sheet = audit_book['GUTS']

    def audit_sheet_sorter (self): #reads all AP's from chosen Audit Sheet and categories them by their color
     ##### global vairables #####
        global site
        global WHID
        global AP_red_list
        global AP_blue_list
        global AP_green_list
        global AP_orange_list
        global AP_grey_list
        global AP_equipment_list
        global AP_height_list
        global AP_broadcasting_list 
        global AP_backboard_list
        global AP_safety_wire_list
        global AP_service_loop_list
        global AP_label_list
        global AP_location_list
        global AP_obstruction_list

        global AP_blue_count_list

        global AP_backboard_count_list
        global AP_broadcasting_count_list
        global AP_service_loop_count_list
        global AP_height_count_list
        global AP_equipment_count_list
        global AP_label_count_list
        global AP_location_count_list
        global AP_obstruction_count_list
        global AP_safety_wire_count_list
        
     ##### COLUMNS #####   
        xlname_col_ID = 'A' #column letter to read AP names from
        xlname_col_equipment = 'B'  #column letter to read equipment fails from
        xlname_col_height = 'C'
        xlname_col_broadcasting = 'D'
        xlname_col_backboard = 'E'
        xlname_col_safety_wire = 'F'
        xlname_col_service_loop = 'G'
        xlname_col_label = 'H'
        xlname_col_location = 'I'
        xlname_col_obstruction = 'J'
        xlname_col_fail = 'K' #column letter to read fails from
        xlname_col_custom_move1_direction = 'M'
        xlname_col_custom_move1_quantity = 'N'
        xlname_col_custom_move1_metric = 'O'
        xlname_col_custom_move2_direction = 'P'
        xlname_col_custom_move2_quantity = 'Q'
        xlname_col_custom_move2_metric = 'R'
        xlname_col_custom_obstruction = 'S'
        xlname_col_custom_name_broadcasting = 'T'
        xlname_col_custom_name_label = 'U'
        xlname_col_custom_fail = 'V'
        xlname_col_custom_fix = 'W'
        xlname_col_section = 'X'
        xlname_col_record_height = 'Y'
     ##### ROWS #####
        xlname_row_ID = int('3') #Starting row in excel
        xlname_row_fail = int('3') #Starting row in excel       
        xlname_row_equipment = int('3') 
        xlname_row_height = int('3')
        xlname_row_broadcasting = int('3')
        xlname_row_backboard = int('3')
        xlname_row_safety_wire = int('3')
        xlname_row_service_loop = int('3')
        xlname_row_label = int('3')
        xlname_row_location = int('3')
        xlname_row_obstruction = int('3')
        results_row = int('2')
     ##### EMPTY LISTS of VARIABLES #####
        site = []  #makes a blank array to hold the SITE string value
        AP_red_list = []  #initiates a list for green AP's ('2')
        AP_blue_list = []  #initiates a list for green AP's ('4')
        AP_green_list = []  #initiates a list for green AP's ('1')
        AP_orange_list = []  #initiates a list for green AP's ('3')
        AP_grey_list = []
        AP_equipment_list = []
        AP_height_list = []
        AP_broadcasting_list =[]
        AP_backboard_list = []
        AP_safety_wire_list = []
        AP_service_loop_list = []
        AP_label_list = []
        AP_location_list = []
        AP_obstruction_list = []
        
     ##### INITIATING COUNTS ##### 
        AP_total_count_list = int('0') 
        AP_red_count_list = int('0')#Keeps track how long the list is
        AP_blue_count_list = int('0')
        AP_green_count_list = int('0')
        AP_orange_count_list = int('0')  
        AP_grey_count_list = int('0')    
        AP_equipment_count_list = int('0')
        AP_height_count_list = int('0')
        AP_broadcasting_count_list = int('0')
        AP_backboard_count_list = int('0')
        AP_safety_wire_count_list = int('0')
        AP_service_loop_count_list = int('0')
        AP_label_count_list = int('0')
        AP_location_count_list = int('0')
        AP_obstruction_count_list = int('0')     
      
     ##### EXCEL EXPORT #####
        global book_results
        global sheet_results
        global sheet_counts
        global sheet_explain
        template = "O:/Customers/Amazon/_Templates/Audit Results Template.xlsx"
        book_results = load_workbook(template)
        #book_results.template = False
        sheet_results = book_results['Results']
        sheet_counts = book_results['Counts'] 
        sheet_explain = book_results['Explanation']
        print('\nAP sorting and data scraping...')

     ##### MAIN for LOOP #####    
        for AP in range (0, audit_sheet.max_row +1): #range of APS being colored
        ##### variable CELLS #####
            xlname_ID_cell = str(xlname_col_ID) + str(xlname_row_ID)
            xlname_fail_cell = str(xlname_col_fail) + str(xlname_row_fail)
            xlname_section_cell = str(xlname_col_section) + str(xlname_row_ID)
            xlname_equipment_cell = str(xlname_col_equipment) + str(xlname_row_equipment)   #equipment
            xlname_height_cell = str(xlname_col_height) + str(xlname_row_height)
            xlname_broadcasting_cell = str(xlname_col_broadcasting) + str(xlname_row_broadcasting)
            xlname_backboard_cell = str(xlname_col_backboard) + str(xlname_row_backboard)
            xlname_saftey_wire_cell = str(xlname_col_safety_wire) + str(xlname_row_safety_wire)
            xlname_service_loop_cell = str(xlname_col_service_loop) + str(xlname_row_service_loop)
            xlname_label_cell = str(xlname_col_label) + str(xlname_row_label)
            xlname_location_cell = str(xlname_col_location) + str(xlname_row_location)
            xlname_obstruction_cell = str(xlname_col_obstruction) + str(xlname_row_obstruction)
            xlname_custom_obstruction_cell = str(xlname_col_custom_obstruction) + str(xlname_row_obstruction)
            xlname_custom_fail_cell = str(xlname_col_custom_fail) + str(xlname_row_ID)
            xlname_custom_fix_cell = str(xlname_col_custom_fix) + str(xlname_row_ID)
            xlname_custom_move1_direction = str(xlname_col_custom_move1_direction) + str(xlname_row_obstruction)
            xlname_custom_move1_quantity = str(xlname_col_custom_move1_quantity) + str(xlname_row_obstruction)
            xlname_custom_move1_metric = str(xlname_col_custom_move1_metric) + str(xlname_row_obstruction)
            xlname_custom_move2_direction = str(xlname_col_custom_move2_direction) + str(xlname_row_obstruction)
            xlname_custom_move2_quantity = str(xlname_col_custom_move2_quantity) + str(xlname_row_obstruction)
            xlname_custom_move2_metric = str(xlname_col_custom_move2_metric) + str(xlname_row_obstruction)
            xlname_custom_name_broadcasting = str(xlname_col_custom_name_broadcasting) + str(xlname_row_broadcasting)
            xlname_custom_name_label = str(xlname_col_custom_name_label) + str(xlname_row_label)
            xlname_record_height = str(xlname_col_record_height) + str(xlname_row_height)
        ##### CELL VALUES #####
            AP_ID = audit_sheet[xlname_ID_cell].value
            AP_fail = audit_sheet[xlname_fail_cell].value
            AP_section = audit_sheet[xlname_section_cell].value
            AP_equipment  = audit_sheet[xlname_equipment_cell].value
            AP_height  = audit_sheet[xlname_height_cell].value
            AP_broadcasting  = audit_sheet[xlname_broadcasting_cell].value
            AP_backboard  = audit_sheet[xlname_backboard_cell].value
            AP_safety_wire  = audit_sheet[xlname_saftey_wire_cell].value
            AP_service_loop  = audit_sheet[xlname_service_loop_cell].value
            AP_label  = audit_sheet[xlname_label_cell].value
            AP_location  = audit_sheet[xlname_location_cell].value
            AP_obstruction = audit_sheet[xlname_obstruction_cell].value
            AP_custom_obstruction = audit_sheet[xlname_custom_obstruction_cell].value
            AP_custom_fail = audit_sheet[xlname_custom_fail_cell].value
            AP_custom_fix = audit_sheet[xlname_custom_fix_cell].value
            AP_custom_move1_direction = audit_sheet[xlname_custom_move1_direction].value
            AP_custom_move1_quantity = audit_sheet[xlname_custom_move1_quantity].value
            AP_custom_move1_metric = audit_sheet[xlname_custom_move1_metric].value
            AP_custom_move2_direction = audit_sheet[xlname_custom_move2_direction].value
            AP_custom_move2_quantity = audit_sheet[xlname_custom_move2_quantity].value
            AP_custom_move2_metric = audit_sheet[xlname_custom_move2_metric].value
            AP_custom_name_broadcasting = audit_sheet[xlname_custom_name_broadcasting].value
            AP_custom_name_label = audit_sheet[xlname_custom_name_label].value
            AP_record_height = audit_sheet[xlname_record_height].value
        ##### SITE NAME #####
            site.append(AP_ID)  #appending cell values of first column to Site
            site_name = site[0] #opens that the first cell in the column
            WHID = site_name[0:4] #read the first 4 characters in cell value (site name)
        
        ##### MOVE DICTIONARIES #####
            move1_direction = {
                "u" : "plan north ",
                "d" : "plan south ",
                "l" : "plan west ",
                "r" : "plan east ",
                "lo" : "lower ",
                "ra" : "raise "
            }
            move2_direction = {
                "u" : " and plan north ",
                "d" : " and plan south ",
                "l" : " and plan west ",
                "r" : " and plan east ",
                "lo" : " and lower ",
                "ra" : " and raise "
            }
            metric_singular = {
                "i" : " inch",
                "f" : " ft", 
                "a" : " aisle", 
                "b" : " bin",
                "c" : " cell",
                "r" : " row",
                "bay" : " bay door",
                "red" : " red iron"  
            }
            metric_plural = {
                "i" : " inches",
                "f" : " ft", 
                "a" : " aisles", 
                "b" : " bins",
                "c" : " cells",
                "r" : " rows",
                "bay" : " bay doors",
                "red" : " red irons" 
            }
        ### direction 1 error handling ###
            try:
                direction1 = move1_direction[str(AP_custom_move1_direction)]
            except KeyError:
                direction1 = str("")
        ### direction 2 error handling ###
            try:
                direction2 = move2_direction[str(AP_custom_move2_direction)]
            except KeyError:
                direction2 = str("") # if teh sports blank, dont need to priont anything
        ### quantity error handling
            if AP_custom_move1_quantity == None:
                AP_custom_move1_quantity = str("")

            if AP_custom_move2_quantity == None:
                AP_custom_move2_quantity = str("")
        ### metric 1 error handling ###
            try:
                if AP_custom_move1_quantity == 1:
                    metric1 = metric_singular[str(AP_custom_move1_metric)]
                elif AP_custom_move1_quantity != 1:
                    metric1 = metric_plural[str(AP_custom_move1_metric)]
            except KeyError:
                metric1 = str("")                        
        ### metric 2 error handling ###
            try:
                if AP_custom_move2_quantity == 1:
                    metric2 = metric_singular[str(AP_custom_move2_metric)]
                elif AP_custom_move2_quantity != 1:
                    metric2 = metric_plural[str(AP_custom_move2_metric)]
            except KeyError:
                metric2 = str("")

        ### custom NONE error handling
            if AP_custom_name_broadcasting == None:
                AP_custom_name_broadcasting = str("")
            if AP_custom_obstruction == None:
                AP_custom_obstruction = str("")
            if AP_custom_name_label == None:
                AP_custom_name_label = str("")

        ##### FAIL DICTIONAIRES #####
            fail_equipment = {
                "1" : guts_sheet['W3'].value,
                "2" : guts_sheet['W4'].value,
                "3" : guts_sheet['W5'].value,
                "4" : guts_sheet['W6'].value,
                "5" : guts_sheet['W7'].value,
                "6" : guts_sheet['W8'].value,
                "7" : guts_sheet['W9'].value,
                "8" : guts_sheet['W10'].value,
                "9" : guts_sheet['W11'].value,
                "10": AP_custom_fail
            }
            fail_height = {
                "1" : guts_sheet['X3'].value,
                "2" : guts_sheet['X4'].value,
                "3" : guts_sheet['X5'].value,
                "4" : guts_sheet['X6'].value,
                "5" : guts_sheet['X7'].value,
                "6" : guts_sheet['X8'].value,
                "7" : guts_sheet['X9'].value,
                "8" : guts_sheet['X10'].value,
                "9" : guts_sheet['X11'].value,
                #"10": AP_custom_fail
                "10": (f"WAP installed at {AP_record_height}")
            }
            fail_broadcasting = {
                "1" : (f"{guts_sheet['Y3'].value}{AP_custom_name_broadcasting}"),
                "2" : (f"{guts_sheet['Y4'].value}{AP_custom_name_broadcasting}"),
                "3" : (f"{guts_sheet['Y5'].value}{AP_custom_name_broadcasting}"),
                "4" : (f"{guts_sheet['Y6'].value}{AP_custom_name_broadcasting}"),
                "5" : (f"{guts_sheet['Y7'].value}{AP_custom_name_broadcasting}"),
                "6" : (f"{guts_sheet['Y8'].value}{AP_custom_name_broadcasting}"),
                "7" : (f"{guts_sheet['Y9'].value}{AP_custom_name_broadcasting}"),
                "8" : (f"{guts_sheet['Y10'].value}{AP_custom_name_broadcasting}"),
                "9" : (f"{guts_sheet['Y11'].value}{AP_custom_name_broadcasting}"),
                "10": AP_custom_fail
            }
            fail_backboard = {
                "1" : guts_sheet['Z3'].value,
                "2" : guts_sheet['Z4'].value,
                "3" : guts_sheet['Z5'].value,
                "4" : guts_sheet['Z6'].value,
                "5" : guts_sheet['Z7'].value,
                "6" : guts_sheet['Z8'].value,
                "7" : guts_sheet['Z9'].value,
                "8" : guts_sheet['Z10'].value,
                "9" : guts_sheet['Z11'].value,
                "10": AP_custom_fail
            }
            fail_safety_wire = {
                "1" : guts_sheet['AA3'].value,
                "2" : guts_sheet['AA4'].value,
                "3" : guts_sheet['AA5'].value,
                "4" : guts_sheet['AA6'].value,
                "5" : guts_sheet['AA7'].value,
                "6" : guts_sheet['AA8'].value,
                "7" : guts_sheet['AA9'].value,
                "8" : guts_sheet['AA10'].value,
                "9" : guts_sheet['AA11'].value,
                "10": AP_custom_fail
            }
            fail_service_loop = {
                "1" : guts_sheet['AB3'].value,
                "2" : guts_sheet['AB4'].value,
                "3" : guts_sheet['AB5'].value,
                "4" : guts_sheet['AB6'].value,
                "5" : guts_sheet['AB7'].value,
                "6" : guts_sheet['AB8'].value,
                "7" : guts_sheet['AB9'].value,
                "8" : guts_sheet['AB10'].value,
                "9" : guts_sheet['AB11'].value,
                "10": AP_custom_fail
            }
            fail_label = {
                "1" : (f"{guts_sheet['AC3'].value}{AP_custom_name_label}"),
                "2" : (f"{guts_sheet['AC4'].value}{AP_custom_name_label}"),
                "3" : (f"{guts_sheet['AC5'].value}{AP_custom_name_label}"),
                "4" : (f"{guts_sheet['AC6'].value}{AP_custom_name_label}"),
                "5" : (f"{guts_sheet['AC7'].value}{AP_custom_name_label}"),
                "6" : (f"{guts_sheet['AC8'].value}{AP_custom_name_label}"),
                "7" : (f"{guts_sheet['AC9'].value}{AP_custom_name_label}"),
                "8" : (f"{guts_sheet['AC10'].value}{AP_custom_name_label}"),
                "9" : (f"{guts_sheet['AC11'].value}{AP_custom_name_label}"),
                "10": AP_custom_fail
            }
            fail_location = {
                "1" : guts_sheet['AD3'].value,
                "2" : guts_sheet['AD4'].value,
                "3" : guts_sheet['AD5'].value,
                "4" : guts_sheet['AD6'].value,
                "5" : guts_sheet['AD7'].value,
                "6" : guts_sheet['AD8'].value,
                "7" : guts_sheet['AD9'].value,
                "8" : guts_sheet['AD10'].value,
                "9" : guts_sheet['AD11'].value,
                "10": AP_custom_fail
            }
            fail_obstruction = {
                "1" : guts_sheet['AE3'].value,
                "2" : (f"{guts_sheet['AE4'].value} {AP_custom_obstruction}"),
                "3" : (f"{guts_sheet['AE5'].value} {AP_custom_obstruction}"),
                "4" : (f"{guts_sheet['AE6'].value} {AP_custom_obstruction}"),
                "5" : (f"{guts_sheet['AE7'].value} {AP_custom_obstruction}"),
                "6" : (f"{guts_sheet['AE8'].value} {AP_custom_obstruction}"),
                "7" : (f"{guts_sheet['AE9'].value} {AP_custom_obstruction}"),
                "8" : (f"{guts_sheet['AE10'].value} {AP_custom_obstruction}"),
                "9" : (f"{guts_sheet['AE11'].value} {AP_custom_obstruction}"),
                "10": AP_custom_fail
            }
        ##### FIX DICTIONARIES #####
            fix_equipment = {
                "1" : guts_sheet['W13'].value,
                "2" : guts_sheet['W14'].value,
                "3" : guts_sheet['W15'].value,
                "4" : guts_sheet['W16'].value,
                "5" : guts_sheet['W17'].value,
                "6" : guts_sheet['W18'].value,
                "7" : guts_sheet['W19'].value,
                "8" : guts_sheet['W20'].value,
                "9" : guts_sheet['W21'].value,
                "10": AP_custom_fix 
            }
            fix_height = {
                "1" : guts_sheet['X13'].value,
                "2" : guts_sheet['X14'].value,
                "3" : guts_sheet['X15'].value,
                "4" : guts_sheet['X16'].value,
                "5" : guts_sheet['X17'].value,
                "6" : guts_sheet['X18'].value,
                "7" : guts_sheet['X19'].value,
                "8" : guts_sheet['X20'].value,
                "9" : guts_sheet['X21'].value,
                "10": AP_custom_fix
            }
            fix_broadcasting = {
                "1" : guts_sheet['Y13'].value,
                "2" : guts_sheet['Y14'].value,
                "3" : guts_sheet['Y15'].value,
                "4" : guts_sheet['Y16'].value,
                "5" : guts_sheet['Y17'].value,
                "6" : guts_sheet['Y18'].value,
                "7" : guts_sheet['Y19'].value,
                "8" : guts_sheet['Y20'].value,
                "9" : guts_sheet['Y21'].value,
                "10": AP_custom_fix
            }
            fix_backboard = {
                "1" : guts_sheet['Z13'].value,
                "2" : guts_sheet['Z14'].value,
                "3" : guts_sheet['Z15'].value,
                "4" : guts_sheet['Z16'].value,
                "5" : guts_sheet['Z17'].value,
                "6" : guts_sheet['Z18'].value,
                "7" : guts_sheet['Z19'].value,
                "8" : guts_sheet['Z20'].value,
                "9" : guts_sheet['Z21'].value,
                "10": AP_custom_fix
            }
            fix_safety_wire = {
                "1" : guts_sheet['AA13'].value,
                "2" : guts_sheet['AA14'].value,
                "3" : guts_sheet['AA15'].value,
                "4" : guts_sheet['AA16'].value,
                "5" : guts_sheet['AA17'].value,
                "6" : guts_sheet['AA18'].value,
                "7" : guts_sheet['AA19'].value,
                "8" : guts_sheet['AA20'].value,
                "9" : guts_sheet['AA21'].value,
                "10": AP_custom_fix
            }
            fix_service_loop = {
                "1" : guts_sheet['AB13'].value,
                "2" : guts_sheet['AB14'].value,
                "3" : guts_sheet['AB15'].value,
                "4" : guts_sheet['AB16'].value,
                "5" : guts_sheet['AB17'].value,
                "6" : guts_sheet['AB18'].value,
                "7" : guts_sheet['AB19'].value,
                "8" : guts_sheet['AB20'].value,
                "9" : guts_sheet['AB21'].value,
                "10": AP_custom_fix
            }
            fix_label = {
                "1" : guts_sheet['AC13'].value,
                "2" : guts_sheet['AC14'].value,
                "3" : guts_sheet['AC15'].value,
                "4" : guts_sheet['AC16'].value,
                "5" : guts_sheet['AC17'].value,
                "6" : guts_sheet['AC18'].value,
                "7" : guts_sheet['AC19'].value,
                "8" : guts_sheet['AC20'].value,
                "9" : guts_sheet['AC21'].value,
                "10": AP_custom_fix
            }
            fix_location = {
                "1" : guts_sheet['AD13'].value,
                "2" : (f"{guts_sheet['AD14'].value}{direction1}{AP_custom_move1_quantity}{metric1}{direction2}{AP_custom_move2_quantity}{metric2}"),
                "3" : (f"{guts_sheet['AD15'].value}{direction1}{AP_custom_move1_quantity}{metric1}{direction2}{AP_custom_move2_quantity}{metric2}"),
                "4" : (f"{guts_sheet['AD16'].value}{direction1}{AP_custom_move1_quantity}{metric1}{direction2}{AP_custom_move2_quantity}{metric2}"),
                "5" : (f"{guts_sheet['AD17'].value}{direction1}{AP_custom_move1_quantity}{metric1}{direction2}{AP_custom_move2_quantity}{metric2}"),
                "6" : (f"{guts_sheet['AD18'].value}{direction1}{AP_custom_move1_quantity}{metric1}{direction2}{AP_custom_move2_quantity}{metric2}"),
                "7" : (f"{guts_sheet['AD19'].value}{direction1}{AP_custom_move1_quantity}{metric1}{direction2}{AP_custom_move2_quantity}{metric2}"),
                "8" : (f"{guts_sheet['AD20'].value}{direction1}{AP_custom_move1_quantity}{metric1}{direction2}{AP_custom_move2_quantity}{metric2}"),
                "9" : (f"{guts_sheet['AD21'].value}{direction1}{AP_custom_move1_quantity}{metric1}{direction2}{AP_custom_move2_quantity}{metric2}"),
                "10": AP_custom_fix
            }
            fix_obstruction = {
                "1" : guts_sheet['AE13'].value,
                "2" : (f"{guts_sheet['AE14'].value}{direction1}{AP_custom_move1_quantity}{metric1}{direction2}{AP_custom_move2_quantity}{metric2}"),
                "3" : (f"{guts_sheet['AE15'].value}{direction1}{AP_custom_move1_quantity}{metric1}{direction2}{AP_custom_move2_quantity}{metric2}"),
                "4" : (f"{guts_sheet['AE16'].value}{direction1}{AP_custom_move1_quantity}{metric1}{direction2}{AP_custom_move2_quantity}{metric2}"),
                "5" : (f"{guts_sheet['AE17'].value}{direction1}{AP_custom_move1_quantity}{metric1}{direction2}{AP_custom_move2_quantity}{metric2}"),
                "6" : (f"{guts_sheet['AE18'].value}{direction1}{AP_custom_move1_quantity}{metric1}{direction2}{AP_custom_move2_quantity}{metric2}"),
                "7" : (f"{guts_sheet['AE19'].value}{direction1}{AP_custom_move1_quantity}{metric1}{direction2}{AP_custom_move2_quantity}{metric2}"),
                "8" : (f"{guts_sheet['AE20'].value}{direction1}{AP_custom_move1_quantity}{metric1}{direction2}{AP_custom_move2_quantity}{metric2}"),
                "9" : (f"{guts_sheet['AE21'].value}{direction1}{AP_custom_move1_quantity}{metric1}{direction2}{AP_custom_move2_quantity}{metric2}"),
                "10": AP_custom_fix
            }
        ##### SEVERITY DICTIONAIRES #####
            severity_equipment = {
                "1" : "No Issue",
                "2" : "LVC",
                "3" : "LVC",
                "4" : "LVC",
                "5" : "LVC",
                "6" : "Attention",
                "7" : "LVC",
                "8" : "LVC",
                "9" : "LVC",
                "10": "CUSTOM"
            }
            severity_height = {
                "1" : "No Issue",
                "2" : "LVC",
                "3" : "LVC",
                "4" : "LVC",
                "5" : "LVC",
                "6" : "LVC",
                "7" : "LVC",
                "8" : "LVC",
                "9" : "LVC",
                "10": "CUSTOM"
            }
            severity_broadcasting = {
                "1" : "No Issue",
                "2" : "IT",
                "3" : "IT",
                "4" : "IT",
                "5" : "IT",
                "6" : "IT",
                "7" : "IT",
                "8" : "IT",
                "9" : "IT",
                "10": "CUSTOM"
            }
            severity_backboard = {
                "1" : "No Issue",
                "2" : "LVC",
                "3" : "LVC",
                "4" : "LVC",
                "5" : "LVC",
                "6" : "LVC",
                "7" : "LVC",
                "8" : "LVC",
                "9" : "LVC",
                "10": "CUSTOM"
            }
            severity_safety_wire = {
                "1" : "No Issue",
                "2" : "LVC",
                "3" : "LVC",
                "4" : "LVC",
                "5" : "LVC",
                "6" : "LVC",
                "7" : "LVC",
                "8" : "LVC",
                "9" : "LVC",
                "10": "CUSTOM"
            }
            severity_service_loop = {
                "1" : "No Issue",
                "2" : "Attention",
                "3" : "LVC",
                "4" : "Attention",
                "5" : "LVC",
                "6" : "LVC",
                "7" : "LVC",
                "8" : "LVC",
                "9" : "LVC",
                "10": "CUSTOM"
            }
            severity_label = {
                "1" : "No Issue",
                "2" : "LVC",
                "3" : "LVC",
                "4" : "LVC",
                "5" : "LVC",
                "6" : "LVC",
                "7" : "LVC",
                "8" : "LVC",
                "9" : "LVC",
                "10": "CUSTOM"
            }
            severity_location = {
                "1" : "No Issue",
                "2" : "LVC",
                "3" : "LVC",
                "4" : "LVC",
                "5" : "LVC",
                "6" : "LVC",
                "7" : "LVC",
                "8" : "LVC",
                "9" : "LVC",
                "10": "CUSTOM"
            }
            severity_obstruction = {
                "1" : "No Issue",
                "2" : "LVC",
                "3" : "LVC",
                "4" : "LVC",
                "5" : "LVC",
                "6" : "LVC",
                "7" : "LVC",
                "8" : "LVC",
                "9" : "LVC",
                "10": "CUSTOM"
            }

        ##### COLOR CHECKS #####   
            if AP_fail == 1: #reads all green AP's
                AP_green_list.append(AP_ID)
                #print(AP_green_list[AP_green_count_list])
                AP_green_count_list += 1
                AP_total_count_list += 1
            elif AP_fail == 2: #reads all red AP's
                AP_red_list.append(AP_ID)
                #print(AP_red_list[AP_red_count_list])
                AP_red_count_list += 1
                AP_total_count_list += 1
                AP_grey_list.append(AP_ID)
                AP_grey_count_list += 1                
            elif AP_fail == 3: #reads all orange AP's
                AP_orange_list.append(AP_ID)
                #print(AP_orange_list[AP_orange_count_list])
                AP_orange_count_list += 1
                AP_total_count_list += 1
                AP_grey_list.append(AP_ID)
                AP_grey_count_list += 1                
            elif AP_fail == 4: #reads all blue AP's
                AP_blue_list.append(AP_ID)
                #print(AP_blue_list[AP_blue_count_list])
                AP_blue_count_list += 1
                AP_total_count_list += 1
                AP_grey_list.append(AP_ID)
                AP_grey_count_list += 1
            elif AP_fail == 5: #reads all red and orange AP's
                AP_red_list.append(AP_ID)
                #print(AP_red_list[AP_red_count_list])
                AP_red_count_list += 1
                AP_orange_list.append(AP_ID)
                #print(AP_orange_list[AP_orange_count_list])
                AP_orange_count_list += 1
                AP_total_count_list += 1
                AP_grey_list.append(AP_ID)
                AP_grey_count_list += 1
            elif AP_fail == 6: #reads all red and blue AP's
                AP_red_list.append(AP_ID)
                #print(AP_red_list[AP_red_count_list])
                AP_red_count_list += 1
                AP_blue_list.append(AP_ID)
                #print(AP_blue_list[AP_blue_count_list])
                AP_blue_count_list += 1
                AP_total_count_list += 1
                AP_grey_list.append(AP_ID)
                AP_grey_count_list += 1
            elif AP_fail == 7: #reads all orange and blue AP's
                AP_orange_list.append(AP_ID)
                #print(AP_orange_list[AP_orange_count_list])
                AP_orange_count_list += 1
                AP_blue_list.append(AP_ID)
                #print(AP_blue_list[AP_blue_count_list])
                AP_blue_count_list += 1
                AP_total_count_list += 1
                AP_grey_list.append(AP_ID)
                AP_grey_count_list += 1
            elif AP_fail == 8: #reads all red and orange AP's
                AP_red_list.append(AP_ID)
                #print(AP_red_list[AP_red_count_list])
                AP_red_count_list += 1
                AP_orange_list.append(AP_ID)
                #print(AP_orange_list[AP_orange_count_list])
                AP_orange_count_list += 1    
                AP_blue_list.append(AP_ID)
                #print(AP_blue_list[AP_blue_count_list])
                AP_blue_count_list += 1
                AP_total_count_list += 1
                AP_grey_list.append(AP_ID)
                AP_grey_count_list += 1
            else:   #at end of Audit Sheet, program exits loop, prints results
                break  
            xlname_row_fail += 1
            xlname_row_ID += 1
        
        ##### SNIP CONTENT #####   Scans Audit sheet for fail numbers, types, severity, issues/fixes, assignes it values in the RESULTS excel
          # equipment 
            if AP_equipment == 1:   #if the number is a passed AP, then porgram passes it
                pass
            elif AP_equipment != 1:              
                AP_equipment_list.append(AP_ID)
                #print(f'{AP_equipment_list[AP_equipment_count_list]} - Equipment #{AP_equipment}    - Problem:{fail_equipment[str(AP_equipment)]} - fix: {fix_equipment[str(AP_equipment)]} - Severity: {severity_equipment[str(AP_equipment)]}')
                results_AP_name_equipment = sheet_results.cell(row = results_row, column = 1, value = (AP_equipment_list[AP_equipment_count_list]))
                results_AP_level_equipment = sheet_results.cell(row = results_row, column = 2, value = AP_section)
                results_issue_category_equipment = sheet_results.cell(row = results_row, column = 3, value = "Equipment")
                results_issue_equipment = sheet_results.cell(row = results_row, column = 4, value = fail_equipment[str(AP_equipment)])
                results_action_equipment = sheet_results.cell(row = results_row, column = 5, value = fix_equipment[str(AP_equipment)]) 
                results_issue_type_equipment = sheet_results.cell(row = results_row, column = 6, value = severity_equipment[str(AP_equipment)])
                results_row += 1  #adds an additional row to the Results Sheet for the next fail to be printed
                AP_equipment_count_list += 1
            xlname_row_equipment += 1 
          # elevation
            if AP_height == 1:
                pass
            elif AP_height != 1:
                AP_height_list.append(AP_ID)
                #print(f'{AP_height_list[AP_height_count_list]} - Height #{AP_height} - Problem:{fail_height[str(AP_height)]} - fix: {fix_height[str(AP_height)]} - Severity: {severity_height[str(AP_height)]}')
                #print(AP_height_list[AP_height_count_list])
                results_AP_name_height = sheet_results.cell(row = results_row, column = 1, value = (AP_height_list[AP_height_count_list]))
                results_AP_level_height = sheet_results.cell(row = results_row, column = 2, value = AP_section)
                results_issue_category_height = sheet_results.cell(row = results_row, column = 3, value = "Elevation")
                results_issue_height = sheet_results.cell(row = results_row, column = 4, value = fail_height[str(AP_height)])
                results_action_height = sheet_results.cell(row = results_row, column = 5, value = fix_height[str(AP_height)]) 
                results_issue_type_height = sheet_results.cell(row = results_row, column = 6, value = severity_height[str(AP_height)])
                results_row += 1
                AP_height_count_list += 1
            xlname_row_height += 1
          # broadcasting
            if AP_broadcasting == 1:
                pass
            elif AP_broadcasting != 1:
                AP_broadcasting_list.append(AP_ID)
                #print(f'{AP_broadcasting_list[AP_broadcasting_count_list]} - Broadcasting #{AP_broadcasting} - Problem:{fail_broadcasting[str(AP_broadcasting)]} - fix: {fix_broadcasting[str(AP_broadcasting)]} - Severity: {severity_broadcasting[str(AP_broadcasting)]}')
                results_AP_name_broadcasting = sheet_results.cell(row = results_row, column = 1, value = (AP_broadcasting_list[AP_broadcasting_count_list]))
                results_AP_level_broadcasting = sheet_results.cell(row = results_row, column = 2, value = AP_section)
                results_issue_category_broadcasting = sheet_results.cell(row = results_row, column = 3, value = "Broadcasting")
                results_issue_broadcasting = sheet_results.cell(row = results_row, column = 4, value = fail_broadcasting[str(AP_broadcasting)])
                results_action_broadcasting = sheet_results.cell(row = results_row, column = 5, value = fix_broadcasting[str(AP_broadcasting)]) 
                results_issue_type_broadcasting = sheet_results.cell(row = results_row, column = 6, value = severity_broadcasting[str(AP_broadcasting)])
                results_row += 1
                AP_broadcasting_count_list += 1
            xlname_row_broadcasting += 1 
          # backboard
            if AP_backboard ==1: 
                pass
            elif AP_backboard!= 1:
                AP_backboard_list.append(AP_ID)
                #print(f'{AP_backboard_list[AP_backboard_count_list]} - Backboard #{AP_backboard}    - Problem:{fail_backboard[str(AP_backboard)]} - fix: {fix_backboard[str(AP_backboard)]} - Severity: {severity_backboard[str(AP_backboard)]}')
                #print(AP_backboard_list[AP_backboard_count_list])
                results_AP_name_backboard = sheet_results.cell(row = results_row, column = 1, value = (AP_backboard_list[AP_backboard_count_list]))
                results_AP_level_backboard = sheet_results.cell(row = results_row, column = 2, value = AP_section)
                results_issue_category_backboard = sheet_results.cell(row = results_row, column = 3, value = "Backboard")
                results_issue_backboard = sheet_results.cell(row = results_row, column = 4, value = fail_backboard[str(AP_backboard)])
                results_action_backboard = sheet_results.cell(row = results_row, column = 5, value = fix_backboard[str(AP_backboard)]) 
                results_issue_type_backboard = sheet_results.cell(row = results_row, column = 6, value = severity_backboard[str(AP_backboard)])
                results_row += 1
                AP_backboard_count_list += 1
            xlname_row_backboard += 1 
          # safety wire
            if AP_safety_wire == 1:
                pass
            elif AP_safety_wire != 1:
                AP_safety_wire_list.append(AP_ID)
                #print(f'{AP_safety_wire_list[AP_safety_wire_count_list]} - Safety Wire #{AP_safety_wire}  - Problem:{fail_safety_wire[str(AP_safety_wire)]} - fix: {fix_safety_wire[str(AP_safety_wire)]} - Severity: {severity_safety_wire[str(AP_safety_wire)]}')
                results_AP_name_safety_wire = sheet_results.cell(row = results_row, column = 1, value = (AP_safety_wire_list[AP_safety_wire_count_list]))
                results_AP_level_safety_wire = sheet_results.cell(row = results_row, column = 2, value = AP_section)
                results_issue_category_safety_wire = sheet_results.cell(row = results_row, column = 3, value = "Safety_Wire")
                results_issue_safety_wire = sheet_results.cell(row = results_row, column = 4, value = fail_safety_wire[str(AP_safety_wire)])
                results_action_safety_wire = sheet_results.cell(row = results_row, column = 5, value = fix_safety_wire[str(AP_safety_wire)]) 
                results_issue_type_safety_wire = sheet_results.cell(row = results_row, column = 6, value = severity_safety_wire[str(AP_safety_wire)])
                results_row += 1
                AP_safety_wire_count_list += 1
            xlname_row_safety_wire += 1 
          # cabling
            if AP_service_loop == 1:
                pass
            elif AP_service_loop != 1:
                AP_service_loop_list.append(AP_ID)
                #print(f'{AP_service_loop_list[AP_service_loop_count_list]} - Serv. Loop #{AP_service_loop}   - Problem:{fail_service_loop[str(AP_service_loop)]} - fix: {fix_service_loop[str(AP_service_loop)]} - Severity: {severity_service_loop[str(AP_service_loop)]}')
                results_AP_name_service_loop = sheet_results.cell(row = results_row, column = 1, value = (AP_service_loop_list[AP_service_loop_count_list]))
                results_AP_level_service_loop = sheet_results.cell(row = results_row, column = 2, value = AP_section)
                results_issue_category_service_loop = sheet_results.cell(row = results_row, column = 3, value = "Cabling")
                results_issue_service_loop = sheet_results.cell(row = results_row, column = 4, value = fail_service_loop[str(AP_service_loop)])
                results_action_service_loop = sheet_results.cell(row = results_row, column = 5, value = fix_service_loop[str(AP_service_loop)]) 
                results_issue_type_service_loop = sheet_results.cell(row = results_row, column = 6, value = severity_service_loop[str(AP_service_loop)])
                results_row += 1
                AP_service_loop_count_list += 1
            xlname_row_service_loop += 1
          # label
            if AP_label ==1:
                pass
            elif AP_label != 1:
                AP_label_list.append(AP_ID)
                #print(f'{AP_label_list[AP_label_count_list]} - Label #{AP_label}        - Problem:{fail_label[str(AP_label)]} - fix:{fix_label[str(AP_label)]} - Severity: {severity_label[str(AP_label)]}')
                results_AP_name_label = sheet_results.cell(row = results_row, column = 1, value = (AP_label_list[AP_label_count_list]))
                results_AP_level_label = sheet_results.cell(row = results_row, column = 2, value = AP_section)
                results_issue_category_label = sheet_results.cell(row = results_row, column = 3, value = "Label")
                results_issue_label = sheet_results.cell(row = results_row, column = 4, value = fail_label[str(AP_label)])
                results_action_label = sheet_results.cell(row = results_row, column = 5, value = fix_label[str(AP_label)]) 
                results_issue_type_label = sheet_results.cell(row = results_row, column = 6, value = severity_label[str(AP_label)])
                results_row += 1
                AP_label_count_list += 1
            xlname_row_label += 1
          # location
            if AP_location == 1:
                pass
            elif AP_location != 1:
                AP_location_list.append(AP_ID)

                #print(f'{AP_location_list[AP_location_count_list]} - Location #{AP_location}     - Problem:{fail_location[str(AP_location)]} - fix:{fix_location[str(AP_location)]} - Severity: {severity_location[str(AP_location)]}')
                results_AP_name_location = sheet_results.cell(row = results_row, column = 1, value = (AP_location_list[AP_location_count_list]))
                results_AP_level_location = sheet_results.cell(row = results_row, column = 2, value = AP_section)
                results_issue_category_location = sheet_results.cell(row = results_row, column = 3, value = "Location")
                results_issue_location = sheet_results.cell(row = results_row, column = 4, value = fail_location[str(AP_location)])
                results_action_location = sheet_results.cell(row = results_row, column = 5, value = fix_location[str(AP_location)]) 
                results_issue_type_location = sheet_results.cell(row = results_row, column = 6, value = severity_location[str(AP_location)])
                results_row += 1
                AP_location_count_list += 1
            xlname_row_location += 1
          # obstruction
            if AP_obstruction == 1:    
                pass
            elif AP_obstruction != 1:
                AP_obstruction_list.append(AP_ID)

                issue_obstruction = fail_obstruction[str(AP_obstruction)]
                action_obstruction = fix_obstruction[str(AP_obstruction)]
                issue_type_obstruction = severity_obstruction[str(AP_obstruction)]

                #print(f'{AP_obstruction_list[AP_obstruction_count_list]} - Obstruction #{AP_obstruction} - Problem:{issue_obstruction} - fix:{action_obstruction} - Severity: {issue_type_obstruction} - Section: {AP_section}')
                 
                results_AP_name_obstruction = sheet_results.cell(row = results_row, column = 1, value = (AP_obstruction_list[AP_obstruction_count_list]))
                results_AP_level_obstruction = sheet_results.cell(row = results_row, column = 2, value = AP_section)
                results_issue_category_obstruction = sheet_results.cell(row = results_row, column = 3, value = "Obstruction")
                results_issue_obstruction = sheet_results.cell(row = results_row, column = 4, value = issue_obstruction)
                results_action_obstruction = sheet_results.cell(row = results_row, column = 5, value = action_obstruction) 
                results_issue_type_obstruction = sheet_results.cell(row = results_row, column = 6, value = issue_type_obstruction)
                results_row += 1

                AP_obstruction_count_list += 1        
            xlname_row_obstruction += 1

     ##### AUDIT RESULTS #####
        print("\n==============================================================================\n")
        print(f'{WHID} Audit Sheet sorted\n')   #stats, used to help user understand the audit
        print(f'    Surveyed APs:       {AP_total_count_list}')
        print(f'    Passed (Green) APs: {AP_green_count_list} ')
        print(f'    Failed (Grey) APs:  {AP_grey_count_list}')
        print(f'     - Red APs:         {AP_red_count_list} ')
        print(f'     - Blue APs:        {AP_blue_count_list} ')
        print(f'     - Orange APs:      {AP_orange_count_list}\n')
        
     ### sheet_counts DATA ###

        #print(f'    Backboard fails:    {AP_backboard_count_list}')
        #print(f'    Broadcasting fails: {AP_broadcasting_count_list}')
        #print(f'    Cabling fails:      {AP_service_loop_count_list}')
        #print(f'    Elevation fails:    {AP_height_count_list}')
        #print(f'    Equipment fails:    {AP_equipment_count_list}')
        #print(f'    Label fails:        {AP_label_count_list}')
        #print(f'    Location fails:     {AP_location_count_list}')
        #print(f'    Obstruction fails:  {AP_obstruction_count_list}')
        #print(f'    Safety Wire fails:  {AP_safety_wire_count_list}\n')
        #print(f'==============================================================================\n')

        #### assigning all the Counts Data in the results excel their respective info
        counts_WHID_value = sheet_counts.cell(row = 2, column = 1, value = WHID)
        #counts_site_type_value = sheet_counts.cell(row = 2, column = 2, value = "**Site Type**")
        #counts_region_value = sheet_counts.cell(row = 2, column = 3, value = "**Region**")
        #counts_survey_type_value = sheet_counts.cell(row = 2, column = 4, value = "**Survey Type**")
        #counts_date_value = sheet_counts.cell(row = 2, column = 5, value = "**Date Performed**")
        counts_AP_total_value = sheet_counts.cell(row = 2, column = 6, value = AP_total_count_list)
        counts_AP_pass_value = sheet_counts.cell(row = 2, column = 7, value = AP_green_count_list)
        counts_AP_fail_value = sheet_counts.cell(row = 2, column = 8, value = AP_grey_count_list)
        #counts_AP_red_value = sheet_counts.cell(row = 2, column = 9, value = AP_red_count_list)
        #counts_AP_orange_value = sheet_counts.cell(row = 2, column = 10, value = AP_orange_count_list)
        #counts_AP_blue_value = sheet_counts.cell(row = 2, column = 11, value = AP_blue_count_list)
        counts_AP_backboard_value = sheet_counts.cell(row = 2, column = 12, value = AP_backboard_count_list)
        counts_AP_broadcast_value = sheet_counts.cell(row = 2, column = 13, value = AP_backboard_count_list)
        counts_AP_service_loop_value = sheet_counts.cell(row = 2, column = 14, value = AP_service_loop_count_list)
        counts_AP_height_value = sheet_counts.cell(row = 2, column = 15, value = AP_height_count_list)
        counts_AP_equipment_value = sheet_counts.cell(row = 2, column = 16, value = AP_equipment_count_list)
        counts_AP_label_value = sheet_counts.cell(row = 2, column = 17, value = AP_label_count_list)
        counts_AP_location_value = sheet_counts.cell(row = 2, column = 18, value = AP_location_count_list)
        ounts_AP_obstruction_value = sheet_counts.cell(row = 2, column = 19, value = AP_obstruction_count_list)
        counts_AP_safety_wire_value = sheet_counts.cell(row = 2, column = 20, value = AP_safety_wire_count_list)
            
        #book_results.save('test_audit_results.xlsx') #saved the information to an excell
        #book_results.save(f'{WHID}.DATE.WAP.SURVEYNAMEResults.CONFIDENTIAL.xlsx')
        #print(f"Exported results to: '{WHID}.DATE.WAP.SURVEYNAMEResults.CONFIDENTIAL.xlsx'\n")
        #exit()

        Options.instructions_starting(self)  #print the main instructions of the program

    def search_options (self):  #sets visio to search all pages for AP IDs, must be done before coloring
        #print (">>> Starting Search...\n")
        print("\nColoring, please standby... \n")
        pyautogui.hotkey('ctrl', 'f')
        pyautogui.press(['tab','tab'])
        pyautogui.press('down')
        pyautogui.press('esc')
        
    def auto_script_grey(self):   #finds all the failed APs, regardless of color 
        AP_grey_count = 0
        while True:
            if keyboard.is_pressed('enter'):     #press enter to start color automation     
                Options.search_options(self)
                for x in AP_grey_list: 
                    AP_num = AP_grey_list[AP_grey_count]
                    AP_num_print = (AP_num[8:13]) if len(AP_num) > 5 else AP_num #Truncates site name, IDF, and AP letter.
                    print(f'     {AP_num_print}')
                    pyautogui.hotkey('ctrl', 'f')
                    pyautogui.typewrite(AP_num_print)
                    pyautogui.press(['enter', 'esc','esc','tab','enter'])
                    pyautogui.hotkey('shift','tab','enter')
                    pyautogui.hotkey('shift','tab','enter')
                    pyautogui.hotkey('alt', 'h')#line coloring
                    pyautogui.press(['l','down','down','down','down','down','left','left','left','enter'])
                    pyautogui.hotkey('alt', 'h')#fill coloring
                    pyautogui.press(['i','down','down','down','down','down','left','left','left','enter'])
                    AP_grey_count += 1     #adds one to total grey count
                    if keyboard.is_pressed('esc'):   #escape sequence if user wants to quit
                        print("Stopping color sequence")
                        break
                    else:   #if esc is not pressed, the for loop will continue iterating
                        pass
                pyautogui.press('esc')
                break
            if keyboard.is_pressed('esc'): #press esc to leave while loops and return to AP Illuminator GUI
                break
        print(f'\nFinished coloring {AP_grey_count} grey APs')
        print("\n==============================================================================\n")
        #Options.instructions_saving(self)

    def auto_script_green(self):  #finds AP name in Visio and colors it GREEN,
        AP_green_count = 0
        while True:
            if keyboard.is_pressed('enter'):
                Options.search_options(self)  
                for x in AP_green_list: 
                    AP_num = AP_green_list[AP_green_count]
                    AP_num_print = (AP_num[8:13]) if len(AP_num) > 5 else AP_num #Truncates site name, IDF, and AP letter.
                    print(f'     {AP_num_print}')
                    pyautogui.hotkey('ctrl', 'f')
                    pyautogui.typewrite(AP_num_print)
                    pyautogui.press(['enter', 'esc','esc','tab','enter'])
                    pyautogui.hotkey('shift','tab','enter')
                    pyautogui.hotkey('shift','tab','enter')
                    pyautogui.hotkey('alt', 'h')#line coloring
                    pyautogui.press(['l','down','down','down','down','down','down','down','left','enter'])
                    pyautogui.hotkey('alt', 'h')#fill coloring
                    pyautogui.press(['i','down','down','down','down','down','down','down','left','enter'])
                    AP_green_count += 1     #adds one to total grey count
                    if keyboard.is_pressed('esc'):
                        print("Stopping color sequence")
                        break
                    else:
                        pass
                pyautogui.press('esc')
                break
            if keyboard.is_pressed('esc'):
                break
        print(f'\nFinished coloring {AP_green_count} green APs')
        print("\n==============================================================================\n")
        #Options.instructions_saving(self)

    def auto_script_red(self): #finds AP name in Visio and colors it RED,   
        AP_red_count = 0 
        while True:
            if keyboard.is_pressed('enter'):
                Options.search_options(self)   
                for x in AP_red_list: #finds AP name in Visio and colors it
                    AP_num = AP_red_list[AP_red_count]
                    AP_num_print = (AP_num[8:13]) if len(AP_num) > 5 else AP_num #Truncates site name, IDF, and AP letter.
                    print(f'     {AP_num_print}')
                    pyautogui.hotkey('ctrl', 'f')
                    pyautogui.typewrite(AP_num_print)
                    pyautogui.press(['enter', 'esc','esc','tab','enter'])
                    pyautogui.hotkey('shift','tab','enter')
                    pyautogui.hotkey('shift','tab','enter')
                    pyautogui.hotkey('alt', 'h')#line coloring
                    pyautogui.press(['l','down','down','down','down','down','down','down','left','left','left','left','enter'])
                    pyautogui.hotkey('alt', 'h')#fill coloring
                    pyautogui.press(['i','down','down','down','down','down','down','down','left','left','left','left','enter'])
                    AP_red_count += 1   #adds one to total red count
                    if keyboard.is_pressed('esc'):
                        print("Stopping color sequence")
                        break
                    else:
                        pass
                pyautogui.press('esc')
                break
            if keyboard.is_pressed('esc'):
                break
        print(f'\nFinished coloring {AP_red_count} red APs')

    def auto_script_orange(self): #finds AP name in Visio and colors it ORANGE, 
        AP_orange_count = 0
        while True:
            if keyboard.is_pressed('enter'):
                Options.search_options(self)
                for x in AP_orange_list: #finds AP name in Visio and colors it
                    AP_num = AP_orange_list[AP_orange_count]
                    AP_num_print = (AP_num[8:13]) if len(AP_num) > 5 else AP_num #Truncates site name, IDF, and AP letter.
                    print(f'     {AP_num_print}')
                    pyautogui.hotkey('ctrl', 'f')
                    pyautogui.typewrite(AP_num_print)
                    pyautogui.press(['enter', 'esc','esc','tab','enter'])
                    pyautogui.hotkey('shift','tab','enter')
                    pyautogui.hotkey('shift','tab','enter')
                    pyautogui.hotkey('alt', 'h')#line coloring
                    pyautogui.press(['l','down','down','down','down','down','down','down','left','left','left','enter'])
                    pyautogui.hotkey('alt', 'h')#fill coloring
                    pyautogui.press(['i','down','down','down','down','down','down','down','left','left','left','enter'])
                    AP_orange_count += 1   # adds one to total orange count
                    if keyboard.is_pressed('esc'):
                        print("Stopping color sequence")
                        break
                    else:
                        pass
                pyautogui.press('esc')
                break
            if keyboard.is_pressed('esc'):
                break
        print(f'\nFinished coloring {AP_orange_count} orange APs')

    def auto_script_blue(self):   #finds AP name in Visio and colors it BLUE,
        AP_blue_count = 0
        while True:
            if keyboard.is_pressed('enter'):
                Options.search_options(self)
                for x in AP_blue_list: 
                    AP_num = AP_blue_list[AP_blue_count]
                    AP_num_print = (AP_num[8:13]) if len(AP_num) > 5 else AP_num #Truncates site name, IDF, and AP letter.
                    print(f'     {AP_num_print}')
                    pyautogui.hotkey('ctrl', 'f')
                    pyautogui.typewrite(AP_num_print)
                    pyautogui.press(['enter', 'esc','esc','tab','enter'])
                    pyautogui.hotkey('shift','tab','enter')
                    pyautogui.hotkey('shift','tab','enter')
                    pyautogui.hotkey('alt', 'h')#line coloring
                    pyautogui.press(['l','down','down','down','down','down','down','down','right','enter'])
                    pyautogui.hotkey('alt', 'h')#fill coloring
                    pyautogui.press(['i','down','down','down','down','down','down','down','right','enter'])
                    AP_blue_count += 1      #adds one to total blue count
                    if keyboard.is_pressed('esc'):
                        print("Stopping color sequence")
                        break
                    #if keyboard.is_pressed('pause'):
                    #   pause()
                    else:
                        pass
                pyautogui.press('esc')
                break
            if keyboard.is_pressed('esc'):
                break          
        print(f'\nFinished coloring {AP_blue_count} blue APs')
          
    def instructions_starting(self): #after sorting through colors and fail types, prints these instructions
        print("INSTRUCTIONS - set up")
        print(" > Duplicate your original Visio Design Sheets, rename: 'Red'")
        print(" > Set all AP markers' text font to WHITE ")
        print(" > Set all border and fill line color to either GREY or GREEN,")
        print("        (depending on the majority, reference above stats)")
        #print("     > Repeat this step for each floor")
        print("\n > Using the pop-up 'Visio AP Illuminator' tool, select either GREEN or GREY,")
        print("        (whichever one you did NOT use for the majority)")
        print("     > Follow the on-screen steps from there\n")
        #print(" > Exit that Visio and duplicate it twice,")
        #print("     > Rename the copies: 'Orange' and 'Blue'\n")
        #print(" > Open any of the new Visio files:")
        #print("     > Use the 'Visio AP Illuminator' tool")
        #print("         > Select the respective color you want to use")
        #print("     > Follow the on-screen steps from there")
        #print("     > Repeat for each Visio and their respective color\n")
        print("==============================================================================\n")
        
    def instructions_copying(self):  #This is not being used in the program, 
        print("\nINSTRUCTIONS - copying")    # use Options.instructions_copying(self) to call it when needed
        print(" > Save and Exit this Visio then duplicate it twice,")
        print("     > Rename the copies: 'Orange' and 'Blue'\n")
        print(" > Open any of the new Visio files:")
        print("     > Use the 'Visio AP Illuminator' tool")
        print("         > Select the respective color you want to use")
        print("     > Follow the on-screen steps from there")
        print("     > Repeat for each Visio and their respective color")
         
    def instructions_saving(self):  #mini while loop, press 'esc' to leave it and return to GUI
        #print(" > To Save as an image for Report Generation, go to:")
        #print("     >  File --> Export --> Change File Type --> SVG (Scalable Vector Graphic)")
        #print("\nINSTRUCTIONS - saving")
        print("\n > Press ' = ' to Save as an SVG image for Report Generation, ")   # '=' 
        print("         > Destination: same as Visio")
        print("         > File Name: 'Color' - 'scope'   Ex. 'Red - A1'")
       # print("     > Repeat for each floor")
        print("\n > Press ' esc ' to return to the 'Visio AP Illuminator' tool")
        print(" > Open another Visio file for a different color or press 'QUIT'\n")
        print("==============================================================================\n")
        while True:    #waiting for user to enter either '=' or 'esc'
            if keyboard.is_pressed('='):  #hotkey for saving visio as image
                #pyautogui.press(['backspace', 'esc'])
                pyautogui.hotkey('alt', 'f')
                pyautogui.press(['e', 'c','g','enter'])
                print("Saving screenshot... \n")
            elif keyboard.is_pressed('esc'):  #exits loop, returns functionality to Color GUI
                break
        #print("Select another color or press 'QUIT'\n")

    def instructions_coloring(self): # the printed instuctions after a color is selected in GUI 
        print(" > Click into your Visio file")
        print(" > Press ' enter ' to start the color sequence")
        print("     > While the program is running,")
        print("           DO NOT click anywhere outide of the Visio file")
        print(" > If you chose the wrong color, press ' esc ' to return to 'AP Illuminator'\n")
        print(" >  Spam ' esc ' while the program is running if you need to exit early  ")
        
    def results_save(self):   #Pop-up message box, asks if user wants to save sorted date to new excel file/sheet
        result = messagebox.askyesno("AP Illuminator","Export data into an Audit Results Template?")
        #print(result)
        if result == True:   #if they chose yes, prints out the fail stats and export destination
            print(f'    Backboard fails:    {AP_backboard_count_list}')
            print(f'    Broadcasting fails: {AP_broadcasting_count_list}')
            print(f'    Cabling fails:      {AP_service_loop_count_list}')
            print(f'    Elevation fails:    {AP_height_count_list}')
            print(f'    Equipment fails:    {AP_equipment_count_list}')
            print(f'    Label fails:        {AP_label_count_list}')
            print(f'    Location fails:     {AP_location_count_list}')
            print(f'    Obstruction fails:  {AP_obstruction_count_list}')
            print(f'    Safety Wire fails:  {AP_safety_wire_count_list}\n')
            #print(f'==============================================================================\n')
            #book_results.save('test_audit_results.xlsx') #saved the information to an excell
            #book_results.save(f"{WHID}.DATE.WAP.SURVEYNAMEResults.CONFIDENTIAL.xlsx")
            today = datetime.date.today()
            x = datetime.datetime(today.year, today.month, today.day)
            #print(x.strftime("%m%d%y"))
            book_results.save(f"O:/Customers/Amazon/{WHID}/Validations/{WHID}.{x.strftime('%m%d%y')}.WAP.SURVEYNAMEResults.CONFIDENTIAL.xlsx")
            print(f"Exported results to: 'O:/Customers/Amazon/{WHID}/Validations/\n    {WHID}.{x.strftime('%m%d%y')}.WAP.SURVEYNAMEResults.CONFIDENTIAL.xlsx'\n")  #export destination
            print(f'==============================================================================\n')
        #exit()  
        else:
            pass

    def results_svg(self):  #dedicated button to open the SVG loop for image saving
        print(" > Click into your Visio file\n")
        print(" > Press ' = ' to Save as an SVG image for Report Generation, ")   # '=' 
        print("         > Destination: same as Visio")
        print("         > File Name: 'Color' - 'scope'   Ex. 'Red - A1'")
       # print("     > Repeat for each floor")
        print("\n > Press ' esc ' to return to the 'Visio AP Illuminator' tool\n")
        print("==============================================================================\n")
        while True:    #waiting for user to enter either '=' or 'esc'
            if keyboard.is_pressed('='):  #hotkey for saving visio as image
                #pyautogui.press(['backspace', 'esc'])
                pyautogui.hotkey('alt', 'f')
                pyautogui.press(['e', 'c','g','enter'])
                print("Saving as SVG... \n")
            elif keyboard.is_pressed('esc'):  #exits loop, returns functionality to Color GUI
                break


class AutoColor(tkinter.Frame):  #POP-UP GUI for choosing majority of Sheets color
    def __init__(self, master=None):  #initializes the GUI
        super().__init__(master)
        self.master = master

        self.grid(column = 0, row = 0)
        self.create_widgets()


    def create_widgets(self):   #buttons in the GUI
        self.title = tkinter.Label(self, font = 30, text = "Select a color to highlight:").grid(column=1, columnspan =3, row =1, pady = 30, ipadx = 30)
       # self.entry_grey = tkinter.Label(self, text = "Page Count: ").grid(column=1, row =2)
       # self.entry_grey = tkinter.Entry(self, width = 5).grid(column=2, row =2)
        self.press_grey = tkinter.Button(self, text = "Grey", fg = "grey", command = self.run_grey).grid(column=2, row =3, ipadx = 8)     
        self.press_green = tkinter.Button(self, text = "Green", fg = "green", command = self.run_green).grid(column=2, row =4, pady = 6, ipadx = 4) 
        self.press_red = tkinter.Button(self, text = "Red", fg = "red", command = self.run_red).grid(column=2, row =5, ipadx = 9) 
        self.press_orange = tkinter.Button(self, text = "Orange", fg = "orange", command = self.run_orange).grid(column=2, row =6, pady = 5)
        self.press_blue = tkinter.Button(self, text = "Blue", fg = "blue", command = self.run_blue).grid(column=2, row =7, ipadx = 8)
        self.save = tkinter.Button(self, text = "Export", command = self.export).grid(column=1, row =8, padx = 20, pady = 20, ipadx = 10) 
        self.quit = tkinter.Button(self, text="QUIT", command=self.close).grid(column=3, row =8, padx = 20, ipadx = 10) 
        self.share = tkinter.Button(self, text = "Save SVG", command = self.svg).grid(column = 2, row = 8) 
        #self.progress = ttk.Progressbar(self, orient="horizontal", length=200, mode="determinate").grid(column=1, columnspan =3, row = 10)
        
    def run_grey(self):   #colors all  green
        print("Color selected: GREY\n")
        Options.instructions_coloring(self)
        Options.auto_script_grey(self) 
        #Options.instructions_copying(self)    
        #Options.instructions_saving(self)

    def run_green(self):
        print("Color selected: GREEN\n")
        Options.instructions_coloring(self)
        Options.auto_script_green(self)
        #Options.instructions_copying(self)
        #Options.instructions_saving(self)

    def run_red(self):  #colors all  red
        print("Color selected: RED\n")
        Options.instructions_coloring(self)
        Options.auto_script_red(self)
        Options.instructions_saving(self)

    def run_orange(self):  #colors all  orange
        print("Color selected: ORANGE\n")
        Options.instructions_coloring(self)
        Options.auto_script_orange(self) 
        Options.instructions_saving(self)    

    def run_blue(self):  #colors all  blue
        print("Color selected: BLUE\n")
        Options.instructions_coloring(self)
        Options.auto_script_blue(self)     
        Options.instructions_saving(self)

    def export(self):  # option to export Stats data into Results report
        Options.results_save(self)

    def svg(self):
        Options.results_svg(self)

    def close(self):  #Ends the porgram when user selcts "QUIT"
        result = messagebox.askyesno("AP Illuminator","Are you sure?")
        if result == True:
            #print("Restarting Program...")
            #start = Options()
            #start.open_audit_sheet()
            #start.audit_sheet_sorter()
            print("Closing program... Have a nice day   :)")
            time.sleep(2)
            #sys.exit()  #for exceutable quitting
            exit()  #for VSC coding
        else:
            #print("Closing program... Have a nice day   :)")
            #time.sleep(1)
            #sys.exit()
            pass
        
        
def main():   #the first loop, initates everything
    keyPress = "" 
    print("Enter '[' to open Visio AP Naming tool")
    print("Enter ']' to open Visio AP Coloring tool")
    #print("Enter 'p' to quit program\n")
    while keyPress != "p": 
        #choice = input(">>> ")
        keyPress = keyboard.read_key()
        #print(f'>>> {keyPress}')
        if keyPress == "[":
            pyautogui.press('backspace')
            print("Starting Visio Naming Tool... \n")           
            theGUI()
            #vnt.theGUI()    #the GUI function calls these internally:
            break               #vnt.startWorkbook()   
                                    #vnt.openFiles()
                                    #vnt.visioLoop()
                                        #vnt.visioGuts()                                      
        elif keyPress == "]":           #vnt.saveExcel()   
            #AutoColor Tool
            pyautogui.press('backspace')
            print("Starting Visio Color Tool... \n")
            start = Options()  #initalizes teh Options Class
            start.open_audit_sheet()  #starts with this module, opens excels
            start.audit_sheet_sorter()
            root = tkinter.Tk()  #gui stuff
            AC = AutoColor(master=root)  #gui stuff
            AC.master.title('AP Illuminator')  #gui stuff
            AC.mainloop()  #gui stuff
            return
        elif keyPress == "p":
            pyautogui.press('backspace')
            print("Closing program... Have a nice day   :)")
            break
        else:
            print("Invalid input")
            pass


if __name__ == "__main__":
    main()   #this is the call to run the whole program



# ******************************
# Changes *********************
# ******************************
    # Adding error handling for NONE types entered
    # prints APs when colored again
    # 
