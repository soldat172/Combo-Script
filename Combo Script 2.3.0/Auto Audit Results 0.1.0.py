import keyboard, openpyxl, pyautogui, tkinter, time, sys, os, datetime
from tkinter import filedialog, messagebox, Tk
from openpyxl import Workbook, load_workbook



class Options:
    def open_audit_sheet (self): #Opens a file explorer and returns path from chosen file
        global audit_sheet
        global guts_sheet
        global audit_book
        audit_book = load_workbook('DPH6 Audit Sheets.xlsx')
        #audit_book = load_workbook('OKC1 Audit Sheets tester.xlsx')

       
        audit_sheet = audit_book['Audit'] #grabs information from the "audit" sheet
        guts_sheet = audit_book['GUTS']

    def audit_sheet_sorter (self): #reads all AP's from chosen Audit Sheet and categories them by their color
    ##### global vairables #####
        global site
        global name
        global AP_red_list
        global AP_blue_list
        global AP_green_list
        global AP_orange_list
        global AP_equipment_list
        global AP_height_list
        global AP_broadcasting_list 
        global AP_backboard_list
        global AP_safety_wire_list
        global AP_service_loop_list
        global AP_label_list
        global AP_location_list
        global AP_obstruction_list
    ##### COLUMNS #####   
        xlname_col_ID = 'A' #column letter to read AP names from
        xlname_col_equipment = 'B'  #column letter to read equipment fails from
        xlname_col_height = 'C'
        xlname_col_broadcasting = 'D'
        xlname_col_backboard = 'E'
        xlname_col_safety_wire = 'F'
        xlname_col_service_loop = 'G'
        xlname_col_label = 'H'
        xlname_col_location = 'I'
        xlname_col_obstruction = 'J'
        xlname_col_fail = 'K' #column letter to read fails from
        xlname_col_custom_move1_direction = 'M'
        xlname_col_custom_move1_quantity = 'N'
        xlname_col_custom_move1_metric = 'O'
        xlname_col_custom_move2_direction = 'P'
        xlname_col_custom_move2_quantity = 'Q'
        xlname_col_custom_move2_metric = 'R'
        xlname_col_custom_obstruction = 'S'
        xlname_col_custom_name_broadcast = 'T'
        xlname_col_custom_name_label = 'U'
        xlname_col_custom_fail = 'V'
        xlname_col_custom_fix = 'W'
        xlname_col_section = 'X'
    ##### ROWS #####
        xlname_row_ID = int('3') #Starting row in excel
        xlname_row_fail = int('3') #Starting row in excel       
        xlname_row_equipment = int('3') 
        xlname_row_height = int('3')
        xlname_row_broadcasting = int('3')
        xlname_row_backboard = int('3')
        xlname_row_safety_wire = int('3')
        xlname_row_service_loop = int('3')
        xlname_row_label = int('3')
        xlname_row_location = int('3')
        xlname_row_obstruction = int('3')
    ##### EMPTY LISTS of VARIABLES #####
        site = []  #makes a blank array to hold the SITE string value
        AP_red_list = []  #initiates a list for green AP's ('2')
        AP_blue_list = []  #initiates a list for green AP's ('4')
        AP_green_list = []  #initiates a list for green AP's ('1')
        AP_orange_list = []  #initiates a list for green AP's ('3')
        AP_equipment_list = []
        AP_height_list = []
        AP_broadcasting_list =[]
        AP_backboard_list = []
        AP_safety_wire_list = []
        AP_service_loop_list = []
        AP_label_list = []
        AP_location_list = []
        AP_obstruction_list = []

        #fail_lvc = int('0')
        #fail_it = int('0')
        #fail_attention = int('0')
    ##### INITIATING COUNTS ##### 
        AP_count_total = int('0') 
        AP_red_count_list = int('0')#Keeps track how long the list is
        AP_blue_count_list = int('0')
        AP_green_count_list = int('0')
        AP_orange_count_list = int('0')      
        AP_equipment_count_list = int('0')
        AP_height_count_list = int('0')
        AP_broadcasting_count_list = int('0')
        AP_backboard_count_list = int('0')
        AP_safety_wire_count_list = int('0')
        AP_service_loop_count_list = int('0')
        AP_label_count_list = int('0')
        AP_location_count_list = int('0')
        AP_obstruction_count_list = int('0')     
    
    ##### EXCEL EXPORT #####
        #wb = load_workbook('results_template.xltx')
        #wb.template = False
        wb = Workbook()
        ws = wb.active
        ws.title = "Explainations"
        ws2 = wb.create_sheet("Results", 1) #creates second sheet, titled "results"
        ws3 = wb.create_sheet("Counts", 2) #creates a third sheet, titled "counts"
        print(f'Sheet titles: {wb.sheetnames}')
        max = ws2.max_row

    ##### MAIN for LOOP #####
        for AP in range (0, audit_sheet.max_row +1): #range of APS being colored
        ##### variable CELLS #####
            xlname_ID_cell = str(xlname_col_ID) + str(xlname_row_ID)
            xlname_fail_cell = str(xlname_col_fail) + str(xlname_row_fail)
            xlname_equipment_cell = str(xlname_col_equipment) + str(xlname_row_equipment)   #equipment
            xlname_height_cell = str(xlname_col_height) + str(xlname_row_height)
            xlname_broadcasting_cell = str(xlname_col_broadcasting) + str(xlname_row_broadcasting)
            xlname_backboard_cell = str(xlname_col_backboard) + str(xlname_row_backboard)
            xlname_saftey_wire_cell = str(xlname_col_safety_wire) + str(xlname_row_safety_wire)
            xlname_service_loop_cell = str(xlname_col_service_loop) + str(xlname_row_service_loop)
            xlname_label_cell = str(xlname_col_label) + str(xlname_row_label)
            xlname_location_cell = str(xlname_col_location) + str(xlname_row_location)
            xlname_obstruction_cell = str(xlname_col_obstruction) + str(xlname_row_obstruction)
            xlname_custom_obstruction_cell = str(xlname_col_custom_obstruction) + str(xlname_row_obstruction)
            xlname_custom_fail_cell = str(xlname_col_custom_fail) + str(xlname_row_obstruction)
            xlname_custom_fix_cell = str(xlname_col_custom_fix) + str(xlname_row_obstruction)
            xlname_custom_move1_direction = str(xlname_col_custom_move1_direction) + str(xlname_row_obstruction)
            xlname_custom_move1_quantity = str(xlname_col_custom_move1_quantity) + str(xlname_row_obstruction)
            xlname_custom_move1_metric = str(xlname_col_custom_move1_metric) + str(xlname_row_obstruction)
            xlname_custom_move2_direction = str(xlname_col_custom_move2_direction) + str(xlname_row_obstruction)
            xlname_custom_move2_quantity = str(xlname_col_custom_move2_quantity) + str(xlname_row_obstruction)
            xlname_custom_move2_metric = str(xlname_col_custom_move2_metric) + str(xlname_row_obstruction)
        ##### CELL VALUES #####
            AP_ID = audit_sheet[xlname_ID_cell].value
            AP_fail = audit_sheet[xlname_fail_cell].value
            AP_equipment  = audit_sheet[xlname_equipment_cell].value
            AP_height  = audit_sheet[xlname_height_cell].value
            AP_broadcasting  = audit_sheet[xlname_broadcasting_cell].value
            AP_backboard  = audit_sheet[xlname_backboard_cell].value
            AP_safety_wire  = audit_sheet[xlname_saftey_wire_cell].value
            AP_service_loop  = audit_sheet[xlname_service_loop_cell].value
            AP_label  = audit_sheet[xlname_label_cell].value
            AP_location  = audit_sheet[xlname_location_cell].value
            AP_obstruction = audit_sheet[xlname_obstruction_cell].value
            AP_custom_obstruction = audit_sheet[xlname_custom_obstruction_cell].value
            AP_custom_fail = audit_sheet[xlname_custom_fail_cell].value
            AP_custom_fix = audit_sheet[xlname_custom_fix_cell].value
            AP_custom_move1_direction = audit_sheet[xlname_custom_move1_direction].value
            AP_custom_move1_quantity = audit_sheet[xlname_custom_move1_quantity].value
            AP_custom_move1_metric = audit_sheet[xlname_custom_move1_metric].value
            AP_custom_move2_direction = audit_sheet[xlname_custom_move2_direction].value
            AP_custom_move2_quantity = audit_sheet[xlname_custom_move2_quantity].value
            AP_custom_move2_metric = audit_sheet[xlname_custom_move2_metric].value

        ##### SITE NAME #####
            site.append(AP_ID)  #appending cell values of first column to Site
            site_name = site[0] #opens that the first cell in the column
            WHID = site_name[0:4] #read the first 4 characters in cell value (site name)
        ##### MOVE DICTIONARIES #####
            move1_direction = {
                "u" : "plan north ",
                "d" : "plan south ",
                "l" : "plan west ",
                "r" : "plan east ",
                "lo" : "lower ",
                "ra" : "raise "
            }
            move2_direction = {
                "u" : "and plan north",
                "d" : "and plan south",
                "l" : "and plan west",
                "r" : "and plan east",
                "lo" : "and lower",
                "ra" : "and raise"
            }
            metric_singular = {
                "i" : "inch",
                "f" : "foot", 
                "a" : "aisle", 
                "b" : "bin",
                "c" : "cell",
                "r" : "row",
                "bay" : "bay door",
                "red" : "red iron"  
            }
            metric_plural = {
                "i" : "inches",
                "f" : "feet", 
                "a" : "aisles", 
                "b" : "bins",
                "c" : "cells",
                "r" : "rows",
                "bay" : "bay doors",
                "red" : "red irons" 
            }
        ### direction 1 error handling ###
            try:
                direction1 = move1_direction[str(AP_custom_move1_direction)]
            except KeyError:
                direction1 = str("*cannot identify direction*")
        ### direction 2 error handling ###
            try:
                direction2 = move2_direction[str(AP_custom_move2_direction)]
            except KeyError:
                direction2 = str('') # if teh sports blank, dont need to priont anything
        ### metric 1 error handling ###
            if AP_custom_move1_metric == 1:
                try:
                    metric1 = metric_singular[str(AP_custom_move1_metric)]
                except KeyError:
                    metric1 = str("*cannot identify metric s 1.2*")
            else:
                try:
                    metric1 = metric_plural[str(AP_custom_move1_metric)]
                except KeyError:
                    metric1 = str("*cannot identify metric p 1.2*")
        ### metric 2 error handling ###
            if AP_custom_move2_metric == 1:
                try:
                    metric2 = metric_singular[str(AP_custom_move2_metric)]
                except KeyError:
                    metric2 = str("*cannot identify metric s 2.2*")
            else:
                try:
                    metric2 = metric_plural[str(AP_custom_move2_metric)]
                except KeyError:
                    metric2 = str("*cannot identify metric p 2.2*")


        ##### FAIL DICTIONAIRES #####
            fail_equipment = {
                "1" : guts_sheet['W3'].value,
                "2" : guts_sheet['W4'].value,
                "3" : guts_sheet['W5'].value,
                "4" : guts_sheet['W6'].value,
                "5" : guts_sheet['W7'].value,
                "6" : guts_sheet['W8'].value,
                "7" : guts_sheet['W9'].value,
                "8" : guts_sheet['W10'].value,
                "9" : guts_sheet['W11'].value,
                "10": guts_sheet['W12'].value
            }
            fail_height = {
                "1" : guts_sheet['X3'].value,
                "2" : guts_sheet['X4'].value,
                "3" : guts_sheet['X5'].value,
                "4" : guts_sheet['X6'].value,
                "5" : guts_sheet['X7'].value,
                "6" : guts_sheet['X8'].value,
                "7" : guts_sheet['X9'].value,
                "8" : guts_sheet['X10'].value,
                "9" : guts_sheet['X11'].value,
                "10": guts_sheet['X12'].value
            }
            fail_broadcasting = {
                "1" : guts_sheet['Y3'].value,
                "2" : guts_sheet['Y4'].value,
                "3" : guts_sheet['Y5'].value,
                "4" : guts_sheet['Y6'].value,
                "5" : guts_sheet['Y7'].value,
                "6" : guts_sheet['Y8'].value,
                "7" : guts_sheet['Y9'].value,
                "8" : guts_sheet['Y10'].value,
                "9" : guts_sheet['Y11'].value,
                "10": guts_sheet['Y12'].value
            }
            fail_backboard = {
                "1" : guts_sheet['Z3'].value,
                "2" : guts_sheet['Z4'].value,
                "3" : guts_sheet['Z5'].value,
                "4" : guts_sheet['Z6'].value,
                "5" : guts_sheet['Z7'].value,
                "6" : guts_sheet['Z8'].value,
                "7" : guts_sheet['Z9'].value,
                "8" : guts_sheet['Z10'].value,
                "9" : guts_sheet['Z11'].value,
                "10": guts_sheet['Z12'].value
            }
            fail_safety_wire = {
                "1" : guts_sheet['AA3'].value,
                "2" : guts_sheet['AA4'].value,
                "3" : guts_sheet['AA5'].value,
                "4" : guts_sheet['AA6'].value,
                "5" : guts_sheet['AA7'].value,
                "6" : guts_sheet['AA8'].value,
                "7" : guts_sheet['AA9'].value,
                "8" : guts_sheet['AA10'].value,
                "9" : guts_sheet['AA11'].value,
                "10": guts_sheet['AA12'].value
            }
            fail_service_loop = {
                "1" : guts_sheet['A3'].value,
                "2" : guts_sheet['AB4'].value,
                "3" : guts_sheet['AB5'].value,
                "4" : guts_sheet['AB6'].value,
                "5" : guts_sheet['AB7'].value,
                "6" : guts_sheet['AB8'].value,
                "7" : guts_sheet['AB9'].value,
                "8" : guts_sheet['AB10'].value,
                "9" : guts_sheet['AB11'].value,
                "10": guts_sheet['AB12'].value
            }
            fail_label = {
                "1" : guts_sheet['AC3'].value,
                "2" : guts_sheet['AC4'].value,
                "3" : guts_sheet['AC5'].value,
                "4" : guts_sheet['AC6'].value,
                "5" : guts_sheet['AC7'].value,
                "6" : guts_sheet['AC8'].value,
                "7" : guts_sheet['AC9'].value,
                "8" : guts_sheet['AC10'].value,
                "9" : guts_sheet['AC11'].value,
                "10": guts_sheet['AC12'].value
            }
            fail_location = {
                "1" : guts_sheet['AD3'].value,
                "2" : guts_sheet['AD4'].value,
                "3" : guts_sheet['AD5'].value,
                "4" : guts_sheet['AD6'].value,
                "5" : guts_sheet['AD7'].value,
                "6" : guts_sheet['AD8'].value,
                "7" : guts_sheet['AD9'].value,
                "8" : guts_sheet['AD10'].value,
                "9" : guts_sheet['AD11'].value,
                "10": guts_sheet['AD12'].value
            }
            fail_obstruction = {
                "1" : guts_sheet['AE3'].value,
                "2" : guts_sheet['AE4'].value,
                "3" : guts_sheet['AE5'].value,
                "4" : guts_sheet['AE6'].value,
                "5" : guts_sheet['AE7'].value,
                "6" : guts_sheet['AE8'].value,
                "7" : guts_sheet['AE9'].value,
                "8" : guts_sheet['AE10'].value,
                "9" : (f"{guts_sheet['AE11'].value}{AP_custom_obstruction}"),
                "10": AP_custom_fail
            }
        ##### FIX DICTIONARIES #####
            fix_equipment = {
                "1" : guts_sheet['W13'].value,
                "2" : guts_sheet['W14'].value,
                "3" : guts_sheet['W15'].value,
                "4" : guts_sheet['W16'].value,
                "5" : guts_sheet['W17'].value,
                "6" : guts_sheet['W18'].value,
                "7" : guts_sheet['W19'].value,
                "8" : guts_sheet['W20'].value,
                "9" : guts_sheet['W21'].value,
                "10": guts_sheet['W22'].value
            }
            fix_height = {
                "1" : guts_sheet['X13'].value,
                "2" : guts_sheet['X14'].value,
                "3" : guts_sheet['X15'].value,
                "4" : guts_sheet['X16'].value,
                "5" : guts_sheet['X17'].value,
                "6" : guts_sheet['X18'].value,
                "7" : guts_sheet['X19'].value,
                "8" : guts_sheet['X20'].value,
                "9" : guts_sheet['X21'].value,
                "10": guts_sheet['X22'].value
            }
            fix_broadcasting = {
                "1" : guts_sheet['Y13'].value,
                "2" : guts_sheet['Y14'].value,
                "3" : guts_sheet['Y15'].value,
                "4" : guts_sheet['Y16'].value,
                "5" : guts_sheet['Y17'].value,
                "6" : guts_sheet['Y18'].value,
                "7" : guts_sheet['Y19'].value,
                "8" : guts_sheet['Y20'].value,
                "9" : guts_sheet['Y21'].value,
                "10": guts_sheet['Y22'].value
            }
            fix_backboard = {
                "1" : guts_sheet['Z13'].value,
                "2" : guts_sheet['Z14'].value,
                "3" : guts_sheet['Z15'].value,
                "4" : guts_sheet['Z16'].value,
                "5" : guts_sheet['Z17'].value,
                "6" : guts_sheet['Z18'].value,
                "7" : guts_sheet['Z19'].value,
                "8" : guts_sheet['Z20'].value,
                "9" : guts_sheet['Z21'].value,
                "10": guts_sheet['Z22'].value
            }
            fix_safety_wire = {
                "1" : guts_sheet['AA13'].value,
                "2" : guts_sheet['AA14'].value,
                "3" : guts_sheet['AA15'].value,
                "4" : guts_sheet['AA16'].value,
                "5" : guts_sheet['AA17'].value,
                "6" : guts_sheet['AA18'].value,
                "7" : guts_sheet['AA19'].value,
                "8" : guts_sheet['AA20'].value,
                "9" : guts_sheet['AA21'].value,
                "10": guts_sheet['AA22'].value
            }
            fix_service_loop = {
                "1" : guts_sheet['AB13'].value,
                "2" : guts_sheet['AB14'].value,
                "3" : guts_sheet['AB15'].value,
                "4" : guts_sheet['AB16'].value,
                "5" : guts_sheet['AB17'].value,
                "6" : guts_sheet['AB18'].value,
                "7" : guts_sheet['AB19'].value,
                "8" : guts_sheet['AB20'].value,
                "9" : guts_sheet['AB21'].value,
                "10": guts_sheet['AB22'].value
            }
            fix_label = {
                "1" : guts_sheet['AC13'].value,
                "2" : guts_sheet['AC14'].value,
                "3" : guts_sheet['AC15'].value,
                "4" : guts_sheet['AC16'].value,
                "5" : guts_sheet['AC17'].value,
                "6" : guts_sheet['AC18'].value,
                "7" : guts_sheet['AC19'].value,
                "8" : guts_sheet['AC20'].value,
                "9" : guts_sheet['AC21'].value,
                "10": guts_sheet['AC22'].value
            }
            fix_location = {
                "1" : guts_sheet['AD13'].value,
                "2" : guts_sheet['AD14'].value,
                "3" : guts_sheet['AD15'].value,
                "4" : guts_sheet['AD16'].value,
                "5" : guts_sheet['AD17'].value,
                "6" : guts_sheet['AD18'].value,
                "7" : guts_sheet['AD19'].value,
                "8" : guts_sheet['AD20'].value,
                "9" : guts_sheet['AD21'].value,
                "10": guts_sheet['AD22'].value
            }
            fix_obstruction = {
                "1" : guts_sheet['AE13'].value,
                "2" : guts_sheet['AE14'].value,
                "3" : guts_sheet['AE15'].value,
                "4" : guts_sheet['AE16'].value,
                "5" : guts_sheet['AE17'].value,
                "6" : guts_sheet['AE18'].value,
                "7" : guts_sheet['AE19'].value,
                "8" : guts_sheet['AE20'].value,
                "9" : (f"{guts_sheet['AE21'].value}{direction1} {AP_custom_move1_quantity} {metric1} {direction2} {AP_custom_move2_quantity} {metric2}"),
                "10": AP_custom_fix
            }
            #{move1_direction[str(AP_custom_move1_direction)]}
            #{move1_metric[str(audit_sheet[xlname_custom_move1_direction].value)]}
        ##### SEVERITY DICTIONAIRES #####
            severity_equipment = {
                "1" : "No Issue",
                "2" : "LVC",
                "3" : "LVC",
                "4" : "LVC",
                "5" : "LVC",
                "6" : "Attention",
                "7" : "LVC",
                "8" : "LVC",
                "9" : "LVC",
                "10": ""
            }
            severity_height = {
                "1" : "No Issue",
                "2" : "LVC",
                "3" : "LVC",
                "4" : "LVC",
                "5" : "LVC",
                "6" : "LVC",
                "7" : "LVC",
                "8" : "LVC",
                "9" : "LVC",
                "10": ""
            }
            severity_broadcasting = {
                "1" : "No Issue",
                "2" : "IT",
                "3" : "IT",
                "4" : "IT",
                "5" : "IT",
                "6" : "IT",
                "7" : "IT",
                "8" : "IT",
                "9" : "IT",
                "10": ""
            }
            severity_backboard = {
                "1" : "No Issue",
                "2" : "LVC",
                "3" : "LVC",
                "4" : "LVC",
                "5" : "LVC",
                "6" : "LVC",
                "7" : "LVC",
                "8" : "LVC",
                "9" : "LVC",
                "10": ""
            }
            severity_safety_wire = {
                "1" : "No Issue",
                "2" : "LVC",
                "3" : "LVC",
                "4" : "LVC",
                "5" : "LVC",
                "6" : "LVC",
                "7" : "LVC",
                "8" : "LVC",
                "9" : "LVC",
                "10": ""
            }
            severity_service_loop = {
                "1" : "No Issue",
                "2" : "Attention",
                "3" : "LVC",
                "4" : "Attention",
                "5" : "LVC",
                "6" : "LVC",
                "7" : "LVC",
                "8" : "LVC",
                "9" : "LVC",
                "10": ""
            }
            severity_label = {
                "1" : "No Issue",
                "2" : "LVC",
                "3" : "LVC",
                "4" : "LVC",
                "5" : "LVC",
                "6" : "LVC",
                "7" : "LVC",
                "8" : "LVC",
                "9" : "LVC",
                "10": ""
            }
            severity_location = {
                "1" : "No Issue",
                "2" : "LVC",
                "3" : "LVC",
                "4" : "LVC",
                "5" : "LVC",
                "6" : "LVC",
                "7" : "LVC",
                "8" : "LVC",
                "9" : "LVC",
                "10": ""
            }
            severity_obstruction = {
                "1" : "No Issue",
                "2" : "LVC",
                "3" : "LVC",
                "4" : "LVC",
                "5" : "LVC",
                "6" : "LVC",
                "7" : "LVC",
                "8" : "LVC",
                "9" : "LVC",
                "10": "CUSTOM"
            }
        ##### COLOR CHECKS #####             
            if AP_fail == 1: #reads all green AP's
                AP_green_list.append(AP_ID)
                #print(AP_green_list[AP_green_count_list])
                AP_green_count_list += 1
                AP_count_total += 1
            elif AP_fail == 2: #reads all red AP's
                AP_red_list.append(AP_ID)
                #print(AP_red_list[AP_red_count_list])
                AP_red_count_list += 1
                AP_count_total += 1 
            elif AP_fail == 3: #reads all orange AP's
                AP_orange_list.append(AP_ID)
                #print(AP_orange_list[AP_orange_count_list])
                AP_orange_count_list += 1
                AP_count_total += 1   
            elif AP_fail == 4: #reads all blue AP's
                AP_blue_list.append(AP_ID)
                #print(AP_blue_list[AP_blue_count_list])
                AP_blue_count_list += 1
                AP_count_total += 1
            elif AP_fail == 5: #reads all red and orange AP's
                AP_red_list.append(AP_ID)
                #print(AP_red_list[AP_red_count_list])
                AP_red_count_list += 1
                AP_orange_list.append(AP_ID)
                #print(AP_orange_list[AP_orange_count_list])
                AP_orange_count_list += 1
                AP_count_total += 1
            elif AP_fail == 6: #reads all red and blue AP's
                AP_red_list.append(AP_ID)
                #print(AP_red_list[AP_red_count_list])
                AP_red_count_list += 1
                AP_blue_list.append(AP_ID)
                #print(AP_blue_list[AP_blue_count_list])
                AP_blue_count_list += 1
                AP_count_total += 1
            elif AP_fail == 7: #reads all orange and blue AP's
                AP_orange_list.append(AP_ID)
                #print(AP_orange_list[AP_orange_count_list])
                AP_orange_count_list += 1
                AP_blue_list.append(AP_ID)
                #print(AP_blue_list[AP_blue_count_list])
                AP_blue_count_list += 1
                AP_count_total += 1
            elif AP_fail == 8: #reads all red and orange AP's
                AP_red_list.append(AP_ID)
                #print(AP_red_list[AP_red_count_list])
                AP_red_count_list += 1
                AP_orange_list.append(AP_ID)
                #print(AP_orange_list[AP_orange_count_list])
                AP_orange_count_list += 1    
                AP_blue_list.append(AP_ID)
                #print(AP_blue_list[AP_blue_count_list])
                AP_blue_count_list += 1
                AP_count_total += 1
            else:   #at end of Audit Sheet, program exits loop, 
                break             
            xlname_row_ID += 1
            xlname_row_fail += 1
        ##### SNIP CONTENT ##### 
            if AP_equipment == 1:
                pass
            elif AP_equipment != 1:                    
                AP_equipment_list.append(AP_ID)
                print(f'{AP_equipment_list[AP_equipment_count_list]} - Equipment #{AP_equipment}   - Problem: {fail_equipment[str(AP_equipment)]} - fix: {fix_equipment[str(AP_equipment)]} - Severity: {severity_equipment[str(AP_equipment)]}')
                AP_equipment_count_list += 1
            xlname_row_equipment += 1 
        
            if AP_height == 1:
                pass
            elif AP_height != 1:
                AP_height_list.append(AP_ID)
                print(f'{AP_height_list[AP_height_count_list]} - Height #{AP_height} - Problem: {fail_height[str(AP_height)]} - fix: {fix_height[str(AP_height)]} - Severity: {severity_height[str(AP_height)]}')
                print(AP_height_list[AP_height_count_list])
                AP_height_count_list += 1
            xlname_row_height += 1

            if AP_broadcasting == 1:
                pass
            elif AP_broadcasting != 1:
                AP_broadcasting_list.append(AP_ID)
                print(f'{AP_broadcasting_list[AP_broadcasting_count_list]} - Broadcasting #{AP_broadcasting} - Problem: {fail_broadcasting[str(AP_broadcasting)]} - fix: {fix_broadcasting[str(AP_broadcasting)]} - Severity: {severity_broadcasting[str(AP_broadcasting)]}')
                AP_broadcasting_count_list += 1
            xlname_row_broadcasting += 1 

            if AP_backboard ==1: 
                pass
            elif AP_backboard!= 1:
                AP_backboard_list.append(AP_ID)
                print(f'{AP_backboard_list[AP_backboard_count_list]} - Backboard #{AP_backboard} - Problem: {fail_backboard[str(AP_backboard)]} - fix: {fix_backboard[str(AP_backboard)]} - Severity: {severity_backboard[str(AP_backboard)]}')
                print(AP_backboard_list[AP_backboard_count_list])
                AP_backboard_count_list += 1
            xlname_row_backboard += 1 

            if AP_safety_wire == 1:
                pass
            elif AP_safety_wire != 1:
                AP_safety_wire_list.append(AP_ID)
                print(f'{AP_safety_wire_list[AP_safety_wire_count_list]} - Safety Wire #{AP_safety_wire} - Problem: {fail_safety_wire[str(AP_safety_wire)]} - fix: {fix_safety_wire[str(AP_safety_wire)]} - Severity: {severity_safety_wire[str(AP_safety_wire)]}')
                AP_saftey_wire_count_list += 1
            xlname_row_safety_wire += 1 

            if AP_service_loop == 1:
                pass
            elif AP_service_loop != 1:
                AP_service_loop_list.append(AP_ID)
                print(f'{AP_service_loop[AP_service_loop_count_list]} - Serv. Loop #{AP_service_loop} - Problem: {fail_service_loop[str(AP_service_loop)]} - fix: {fix_service_loop[str(AP_service_loop)]} - Severity: {severity_service_loop[str(AP_service_loop)]}')
                AP_service_loop_count_list += 1
            xlname_row_service_loop += 1

            if AP_label ==1:
                pass
            elif AP_label != 1:
                AP_label_list.append(AP_ID)
                print(f'{AP_label_list[AP_label_count_list]} - Label #{AP_label}       - Problem:{fail_label[str(AP_label)]} - fix:{fix_label[str(AP_label)]} - Severity: {severity_label[str(AP_label)]}')
                AP_label_count_list += 1
            xlname_row_label += 1

            if AP_location == 1:
                pass
            elif AP_location != 1:
                AP_location_list.append(AP_ID)
                print(f'{AP_location_list[AP_location_count_list]} - Location #{AP_location} - Problem:{fail_location[str(AP_location)]} - fix:{fix_location[str(AP_location)]} - Severity: {severity_location[str(AP_location)]}')
                AP_location_count_list += 1
            xlname_row_location += 1

            if AP_obstruction == 1:    
                pass
            elif AP_obstruction != 1:
                AP_obstruction_list.append(AP_ID)

                issue_obstruction = fail_obstruction[str(AP_obstruction)]
                action_obstruction = fix_obstruction[str(AP_obstruction)]
                issue_type_obstruction = severity_obstruction[str(AP_obstruction)]

                print(f'{AP_obstruction_list[AP_obstruction_count_list]} - Obstruction #{AP_obstruction} - Problem:{issue_obstruction} - fix:{action_obstruction} - Severity: {severity_obstruction[str(AP_obstruction)]}')
                
                results_AP_name_obstruction = ws2.cell(row = 1 + max, column = 1, value = (AP_obstruction_list[AP_obstruction_count_list]))
                results_issue_category_obstruction = ws2.cell(row = 1 + max, column = 3, value = "Obstruction")
                results_issue_obstruction = ws2.cell(row = 1 + max, column = 4, value = issue_obstruction)
                results_action_obstruction = ws2.cell(row = 1 + max, column = 5, value = action_obstruction) 
                results_issue_type_obstruction = ws2.cell(row = 1 + max, column = 6, value = issue_type_obstruction)


                AP_obstruction_count_list += 1        
            xlname_row_obstruction += 1
    ##### AUDIT RESULTS #####
        print(f'\n{WHID} Audit Sheet sorted\n')    
        print(f'Surveyed APs:       {AP_count_total}')
        print(f'Passed (Green) APs: {AP_green_count_list} ')
        print(f'Failed APs:         {AP_count_total - AP_green_count_list}')
        print(f'Red APs:            {AP_red_count_list} ')
        print(f'Blue APs:           {AP_blue_count_list} ')
        print(f'Orange APs:         {AP_orange_count_list}\n')
        print(f'Equipment fails:    {AP_equipment_count_list}')
        print(f'Height fails:       {AP_height_count_list}')
        print(f'Broadcasting fails: {AP_broadcasting_count_list}')
        print(f'Backboard fails:    {AP_backboard_count_list}')
        print(f'Safety Wire fails:  {AP_safety_wire_count_list}')
        print(f'Service Loop fails: {AP_service_loop_count_list}')
        print(f'Label fails:        {AP_label_count_list}')
        print(f'Location fails:     {AP_location_count_list}')
        print(f'Obstruction fails:  {AP_obstruction_count_list}\n')
    
    ### sheets_results Titles ###
        results_AP_name = ws2.cell(row = 1, column = 1, value = "AP Name")
        results_section = ws2.cell(row = 1, column = 2, value = "Section")
        results_issue_category = ws2.cell(row = 1, column = 3, value = "Issue Category")
        results_issue = ws2.cell(row = 1, column = 4, value = "Issue")
        results_action = ws2.cell(row = 1, column = 5, value = "Action")
        results_issue_type = ws2.cell(row = 1, column = 6, value = "Issue Type")
    ### sheets_results DATA ###
    

    ### sheet_counts Titles ###
        counts_WHID = ws3.cell(row = 1, column = 1, value = "WHID")
        counts_site_type = ws3.cell(row = 1, column = 2, value = "Site Type")
        counts_region = ws3.cell(row = 1, column = 3, value = "Region")
        counts_survey_type = ws3.cell(row = 1, column = 4, value = "Survey Type")
        counts_date = ws3.cell(row = 1, column = 5, value = "Date Performed")
        counts_AP_total = ws3.cell(row = 1, column = 6, value = "Total APs Audited")
        counts_AP_pass = ws3.cell(row = 1, column = 7, value = "No Issues")
        counts_AP_fail = ws3.cell(row = 1, column = 8, value = "With Issues")
        counts_AP_red = ws3.cell(row = 1, column = 9, value = "LVC Issues")
        counts_AP_orange = ws3.cell(row = 1, column = 10, value = "IT Issues")
        counts_AP_blue = ws3.cell(row = 1, column = 11, value = "Attention Issues")
        counts_AP_backboard = ws3.cell(row = 1, column = 12, value = "Backboard")
        counts_AP_broadcast = ws3.cell(row = 1, column = 13, value = "Broadcast")
        counts_AP_service_loop = ws3.cell(row = 1, column = 14, value = "Cabling")
        counts_AP_height = ws3.cell(row = 1, column = 15, value = "Elevation")
        counts_AP_equipment = ws3.cell(row = 1, column = 16, value = "Equipment")
        counts_AP_label = ws3.cell(row = 1, column = 17, value = "Label")
        counts_AP_location = ws3.cell(row = 1, column = 18, value = "Location")
        counts_AP_obstruction = ws3.cell(row = 1, column = 19, value = "Obstruction")
        counts_AP_safety_wire = ws3.cell(row = 1, column = 20, value = "Safety Wire")
    ### sheet_counts DATA ###
        counts_WHID_value = ws3.cell(row = 2, column = 1, value = WHID)
        counts_site_type_value = ws3.cell(row = 2, column = 2, value = "**Site Type**")
        counts_region_value = ws3.cell(row = 2, column = 3, value = "**Region**")
        counts_survey_type_value = ws3.cell(row = 2, column = 4, value = "**Survey Type**")
        counts_date_value = ws3.cell(row = 2, column = 5, value = "**Date Performed**")
        counts_AP_total_value = ws3.cell(row = 2, column = 6, value = AP_count_total)
        counts_AP_pass_value = ws3.cell(row = 2, column = 7, value = AP_green_count_list)
        counts_AP_fail_value = ws3.cell(row = 2, column = 8, value = AP_count_total - AP_green_count_list)
        counts_AP_red_value = ws3.cell(row = 2, column = 9, value = AP_red_count_list)
        counts_AP_orange_value = ws3.cell(row = 2, column = 10, value = AP_orange_count_list)
        counts_AP_blue_value = ws3.cell(row = 2, column = 11, value = AP_blue_count_list)
        counts_AP_backboard_value = ws3.cell(row = 2, column = 12, value = AP_backboard_count_list)
        counts_AP_broadcast_value = ws3.cell(row = 2, column = 13, value = AP_backboard_count_list)
        counts_AP_service_loop_value = ws3.cell(row = 2, column = 14, value = AP_service_loop_count_list)
        counts_AP_height_value = ws3.cell(row = 2, column = 15, value = AP_height_count_list)
        counts_AP_equipment_value = ws3.cell(row = 2, column = 16, value = AP_equipment_count_list)
        counts_AP_label_value = ws3.cell(row = 2, column = 17, value = AP_label_count_list)
        counts_AP_location_value = ws3.cell(row = 2, column = 18, value = AP_location_count_list)
        counts_AP_obstruction_value = ws3.cell(row = 2, column = 19, value = AP_obstruction_count_list)
        counts_AP_safety_wire_value = ws3.cell(row = 2, column = 20, value = AP_safety_wire_count_list)
            
        wb.save('test_audit_results.xlsx') #saved the information to an excell
        print("Exported results to: test_audit_results.xlsx")


##### VISIO AUTO SCRIPTS #####
    def search_options (self):  #sets visio to search all pages for AP IDs, must be done before coloring
        print ("\nStarting Search...")
        pyautogui.hotkey('ctrl', 'f')
        pyautogui.press(['tab','tab'])
        pyautogui.press('down')
        pyautogui.press('esc')
        
    def auto_script_grey(self):
        # cntr + page down/up to navigate visio pages       
        pyautogui.hotkey('ctrl','a')  #Coloring all AP GREY
        pyautogui.hotkey('alt', 'h')#line coloring
        pyautogui.press(['l','down','down','down','down','down','left','left','left','enter'])
        pyautogui.hotkey('alt', 'h')#fill coloring
        pyautogui.press(['i','down','down','down','down','down','left','left','left','enter'])
        pyautogui.hotkey('alt', 'h')   #Font coloring to white
        pyautogui.hotkey('f', 'c')   
        pyautogui.press(['down','left','left','left','left','left','enter'])
        print("All APs colored to GREY, font set to WHITE")

    def auto_script_green(self):  #finds AP name in Visio and colors it GREEN,
        Options.search_options(self)
        AP_green_count = 0
        for x in AP_green_list: 
            AP_num = AP_green_list[AP_green_count]
            AP_num_print = (AP_num[8:13]) if len(AP_num) > 5 else AP_num #Truncates site name, IDF, and AP letter.
            print(AP_num_print)
            pyautogui.hotkey('ctrl', 'f')
            pyautogui.typewrite(AP_num_print)
            pyautogui.press(['enter', 'esc','esc','tab','enter'])
            pyautogui.hotkey('shift','tab','enter')
            pyautogui.hotkey('shift','tab','enter')
            pyautogui.hotkey('alt', 'h')#line coloring
            pyautogui.press(['l','down','down','down','down','down','down','down','enter'])
            pyautogui.hotkey('alt', 'h')#fill coloring
            pyautogui.press(['i','down','down','down','down','down','down','down','enter'])
            AP_green_count += 1     #adds one to total grey count
        print(f'Colored {AP_green_count} green APs')
        print("Select another color or press 'QUIT'\n")

    def auto_script_red(self): #finds AP name in Visio and colors it RED, 
        Options.search_options(self)    
        AP_red_count = 0
        for x in AP_red_list: #finds AP name in Visio and colors it
            AP_num = AP_red_list[AP_red_count]
            AP_num_print = (AP_num[8:13]) if len(AP_num) > 5 else AP_num #Truncates site name, IDF, and AP letter.
            print(AP_num_print)
            pyautogui.hotkey('ctrl', 'f')
            pyautogui.typewrite(AP_num_print)
            pyautogui.press(['enter', 'esc','esc','tab','enter'])
            pyautogui.hotkey('shift','tab','enter')
            pyautogui.hotkey('shift','tab','enter')
            pyautogui.hotkey('alt', 'h')#line coloring
            pyautogui.press(['l','down','down','down','down','down','down','down','left','left','left','left','enter'])
            pyautogui.hotkey('alt', 'h')#fill coloring
            pyautogui.press(['i','down','down','down','down','down','down','down','left','left','left','left','enter'])
            #Options.font_white(self)
            AP_red_count += 1   #adds one to total red count
        print(f'Colored {AP_red_count} red APs')
        print("Select another color or press 'QUIT'\n")

    def auto_script_orange(self): #finds AP name in Visio and colors it ORANGE,
        Options.search_options(self)
        AP_orange_count = 0
        for x in AP_orange_list: #finds AP name in Visio and colors it
            AP_num = AP_orange_list[AP_orange_count]
            AP_num_print = (AP_num[8:13]) if len(AP_num) > 5 else AP_num #Truncates site name, IDF, and AP letter.
            print(AP_num_print)
            pyautogui.hotkey('ctrl', 'f')
            pyautogui.typewrite(AP_num_print)
            pyautogui.press(['enter', 'esc','esc','tab','enter'])
            pyautogui.hotkey('shift','tab','enter')
            pyautogui.hotkey('shift','tab','enter')
            pyautogui.hotkey('alt', 'h')#line coloring
            pyautogui.press(['l','down','down','down','down','down','down','down','left','left','left','enter'])
            pyautogui.hotkey('alt', 'h')#fill coloring
            pyautogui.press(['i','down','down','down','down','down','down','down','left','left','left','enter'])
            #Options.font_white(self)
            AP_orange_count += 1   # adds one to total orange count
        print(f'Colored {AP_orange_count} orange APs')
        print("Select another color or press 'QUIT'\n")

    def auto_script_blue(self):   #finds AP name in Visio and colors it BLUE,
        Options.search_options(self)
        AP_blue_count = 0
        for x in AP_blue_list: 
            AP_num = AP_blue_list[AP_blue_count]
            AP_num_print = (AP_num[8:13]) if len(AP_num) > 5 else AP_num #Truncates site name, IDF, and AP letter.
            print(AP_num_print)
            pyautogui.hotkey('ctrl', 'f')
            pyautogui.typewrite(AP_num_print)
            pyautogui.press(['enter', 'esc','esc','tab','enter'])
            pyautogui.hotkey('shift','tab','enter')
            pyautogui.hotkey('shift','tab','enter')
            pyautogui.hotkey('alt', 'h')#line coloring
            pyautogui.press(['l','down','down','down','down','down','down','down','right','enter'])
            pyautogui.hotkey('alt', 'h')#fill coloring
            pyautogui.press(['i','down','down','down','down','down','down','down','right','enter'])
            #Options.font_white(self)
            AP_blue_count += 1      #adds one to total blue count
        print(f'Colored {AP_blue_count} blue APs')
        print("Select another color or press 'QUIT'\n")
    

class AutoColor(tkinter.Frame):  #POP-UP GUI for choosing colors
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.grid(column = 0, row = 0)
        self.create_widgets()

    def create_widgets(self):   #buttons in the GUI
        self.title = tkinter.Label(self, font = 30, text = "Select a color to highlight:").grid(column=2, row =1, pady = 30, ipadx = 30)

       # self.entry_grey = tkinter.Label(self, text = "Page Count: ").grid(column=1, row =2)
       # self.entry_grey = tkinter.Entry(self, width = 5).grid(column=2, row =2)
        self.press_grey = tkinter.Button(self, fg = "grey", text = "GREY", command = self.run_grey).grid(column=2, row =3, pady = 10)     
        self.press_green = tkinter.Button(self, text = "Green", fg = "green", command = self.run_green).grid(column=2, row =4) 
        self.press_red = tkinter.Button(self, text = "Red", fg = "red", command = self.run_red).grid(column=2, row =5) 
        self.press_orange = tkinter.Button(self, text = "Orange", fg = "orange", command = self.run_orange).grid(column=2, row =6)
        self.press_blue = tkinter.Button(self, text = "Blue", fg = "blue", command = self.run_blue).grid(column=2, row =7)
        self.quit = tkinter.Button(self, text="QUIT", command=self.close).grid(column=2, row =8, pady = 30) 
      

    def run_grey(self):   #colors all  green
        print("Setting all APs to Grey with WHITE font")
        print("Please double-click into your Visio file.\n")
        time.sleep(3)
        Options.auto_script_grey(self)     
        
    def run_green(self):
        print("Coloring all GREEN APs...")
        print("Please double-click into your Visio file.\n")
        time.sleep(3)
        Options.auto_script_green(self)

    def run_red(self):  #colors all  red
        print("Coloring all RED APs...")
        print("Please double-click into your Visio file.\n")
        time.sleep(3)
        Options.auto_script_red(self)

    def run_orange(self):  #colors all  orange
        print("Coloring all ORANGE APs...")
        print("Please double-click into your Visio file.\n")
        time.sleep(3)   
        Options.auto_script_orange(self)     

    def run_blue(self):  #colors all  blue
        print("Coloring all BLUE APs...")
        print("Please double-click into your Visio file.\n")
        time.sleep(3)
        Options.auto_script_blue(self)     
    
    def close(self):  #Ends the porgram when user selcts "QUIT"
        result = messagebox.askyesno("Visio AP Coloring Tool","Do you want to color a new audit?")
        if result == True:
            print("Restarting Program...")
            start = Options()
            start.open_audit_sheet()
            start.audit_sheet_sorter()
        else:
            print("Closing program... Have a nice day   :)")
            time.sleep(1)
            sys.exit()
        
        
def main():
    keyPress = "" 
    print("Enter '[' to open Visio AP Naming tool")
    print("Enter ']' to open Visio AP Coloring tool")
    print("Enter 'p' to quit program\n")
    while keyPress != "p": 
        #choice = input(">>> ")
        keyPress = keyboard.read_key()
        print(f'>>> {keyPress}')
        if keyPress == "[":
            pass
                                        #vnt.visioGuts()                                      
        elif keyPress == "]":           #vnt.saveExcel()   
            #AutoColor Tool
            pyautogui.press('backspace')
            print("Starting Visio Color Tool\n")
            start = Options()
            start.open_audit_sheet()
            start.audit_sheet_sorter()
            #start.open_guts_sheet()
            root = tkinter.Tk()
            AC = AutoColor(master=root)
            AC.master.title('V.C.T.')
            AC.mainloop()
            return
        elif keyPress == "p":
            pyautogui.press('backspace')
            print("Closing program... Have a nice day   :)")
            break
        else:
            print("Invalid input")
            pass


if __name__ == "__main__":
    main()   #this is the call to run the whole program



# ******************************
# Changes *********************
# ******************************
# added functionality to asking user if they would like to restartthe program
#     with a new audit sheet rather than quiting and opening program again
# Changed Color GUI to Grid layout instead of Pack layout
# Added abiliity to read site Name
# 
