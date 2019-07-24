import keyboard, openpyxl, pyautogui, tkinter, time
from tkinter import filedialog, messagebox, Tk
from openpyxl import Workbook, load_workbook

class Options:
    def open_audit_sheet (self): #Opens a file explorer and returns path from chosen file
        global audit_sheet
        Tk().withdraw() # We don't want a full GUI, so keep the root window from appearing
        xlname = filedialog.askopenfilename() # Show an "Open" dialog box and return the path to the selected file
        print("Loading Excell File... ")
        audit_book = load_workbook(xlname)#Opens Excel
        audit_sheet = audit_book['Audit'] #grabs information from the "audit" sheet

    def audit_sheet_sorter (self): #reads all AP's from chosen Audit Sheet and categories them by their color
        global AP_red_list
        global AP_blue_list
        global AP_green_list
        global AP_orange_list
        
        xlname_col = 'A' #column letter to read AP names from
        xlname_col_fail = 'K' #column letter to read fails from
        xlname_row = int('3') #Starting row in excel
        xlname_row_fail = int('3') #Starting row in excel
        
        AP_red_list = []  #initiates a list for green AP's ('2')
        AP_blue_list = []  #initiates a list for green AP's ('4')
        AP_green_list = []  #initiates a list for green AP's ('1')
        AP_orange_list = []  #initiates a list for green AP's ('3')
        
        AP_red_count_list = int('0')#Keeps track how long the list is
        AP_blue_count_list = int('0')
        AP_green_count_list = int('0')
        AP_orange_count_list = int('0')
        
        print("List of Failed AP ID's:")
        for AP in range (0,15): #range of APS being colored
            xlname_col_fail_num = str(xlname_col_fail) + str(xlname_row_fail)
            xlname_col_num = str(xlname_col) + str(xlname_row)
            cell_value1 = audit_sheet[xlname_col_fail_num]
            cell_value2 = audit_sheet[xlname_col_num]
            AP_fail = cell_value1.value
            AP_ID = cell_value2.value
            
            if AP_fail == 1: #reads all green AP's
                AP_green_list.append(AP_ID)
                print(AP_green_list[AP_green_count_list])
                AP_green_count_list += 1

            elif AP_fail == 2: #reads all red AP's
                AP_red_list.append(AP_ID)
                print(AP_red_list[AP_red_count_list])
                AP_red_count_list += 1
                
            elif AP_fail == 3: #reads all orange AP's
                AP_orange_list.append(AP_ID)
                print(AP_orange_list[AP_orange_count_list])
                AP_orange_count_list += 1
                
            elif AP_fail == 4: #reads all blue AP's
                AP_blue_list.append(AP_ID)
                print(AP_blue_list[AP_blue_count_list])
                AP_blue_count_list += 1

            else:
                print('End of Audit Sheets')
                break   
            xlname_row_fail += 1
            xlname_row += 1

    def search_options (self):  #sets visio to search all pages for AP IDs, must be done before coloring
        print ("\nStarting Search...")
        pyautogui.hotkey('ctrl', 'f')
        pyautogui.press(['tab','tab'])
        pyautogui.press('down')
        pyautogui.press('esc')

    def auto_script_green(self):  #finds AP name in Visio and colors it GREEN,
        AP_green_count = 0
        for x in AP_green_list: 
            AP_num = AP_green_list[AP_green_count]
            AP_num_print = (AP_num[8:13]) if len(AP_num) > 5 else AP_num #Truncates site name, IDF, and AP letter.
            print(AP_num_print)
            pyautogui.hotkey('ctrl', 'f')
            pyautogui.typewrite(AP_num_print)
            pyautogui.press(['enter', 'esc','esc','tab','enter'])
            pyautogui.hotkey('shift','tab','enter')
            pyautogui.hotkey('shift','tab','enter')
            pyautogui.hotkey('alt', 'h')#line coloring
            pyautogui.press(['l','down','down','down','down','down','down','down','left','enter'])
            pyautogui.hotkey('alt', 'h')#fill coloring
            pyautogui.press(['i','down','down','down','down','down','down','down','left','enter'])
            AP_green_count += 1     #adds one to total green count
        print(f'Green APs: {AP_green_count}')

    def auto_script_red(self):    #finds AP name in Visio and colors it RED, 
        AP_red_count = 0
        for x in AP_red_list: #finds AP name in Visio and colors it
            AP_num = AP_red_list[AP_red_count]
            AP_num_print = (AP_num[8:13]) if len(AP_num) > 5 else AP_num #Truncates site name, IDF, and AP letter.
            print(AP_num_print)
            pyautogui.hotkey('ctrl', 'f')
            pyautogui.typewrite(AP_num_print)
            pyautogui.press(['enter', 'esc','esc','tab','enter'])
            pyautogui.hotkey('shift','tab','enter')
            pyautogui.hotkey('shift','tab','enter')
            pyautogui.hotkey('alt', 'h')#line coloring
            pyautogui.press(['l','down','down','down','down','down','down','down','left','left','left','left','enter'])
            pyautogui.hotkey('alt', 'h')#fill coloring
            pyautogui.press(['i','down','down','down','down','down','down','down','left','left','left','left','enter'])
            AP_red_count += 1   #adds one to total red count
        print(f'Red APs: {AP_red_count}')

    def auto_script_orange(self): #finds AP name in Visio and colors it ORANGE,
        AP_orange_count = 0
        for x in AP_orange_list: #finds AP name in Visio and colors it
            AP_num = AP_orange_list[AP_orange_count]
            AP_num_print = (AP_num[8:13]) if len(AP_num) > 5 else AP_num #Truncates site name, IDF, and AP letter.
            print(AP_num_print)
            pyautogui.hotkey('ctrl', 'f')
            pyautogui.typewrite(AP_num_print)
            pyautogui.press(['enter', 'esc','esc','tab','enter'])
            pyautogui.hotkey('shift','tab','enter')
            pyautogui.hotkey('shift','tab','enter')
            pyautogui.hotkey('alt', 'h')#line coloring
            pyautogui.press(['l','down','down','down','down','down','down','down','left','left','left','enter'])
            pyautogui.hotkey('alt', 'h')#fill coloring
            pyautogui.press(['i','down','down','down','down','down','down','down','left','left','left','enter'])
            AP_orange_count += 1   # adds one to total orange count
        print(f'Orange APs: {AP_orange_count}')

    def auto_script_blue(self):   #finds AP name in Visio and colors it BLUE,
        AP_blue_count = 0
        for x in AP_blue_list: 
            AP_num = AP_blue_list[AP_blue_count]
            AP_num_print = (AP_num[8:13]) if len(AP_num) > 5 else AP_num #Truncates site name, IDF, and AP letter.
            print(AP_num_print)
            pyautogui.hotkey('ctrl', 'f')
            pyautogui.typewrite(AP_num_print)
            pyautogui.press(['enter', 'esc','esc','tab','enter'])
            pyautogui.hotkey('shift','tab','enter')
            pyautogui.hotkey('shift','tab','enter')
            pyautogui.hotkey('alt', 'h')#line coloring
            pyautogui.press(['l','down','down','down','down','down','down','down','right','enter'])
            pyautogui.hotkey('alt', 'h')#fill coloring
            pyautogui.press(['i','down','down','down','down','down','down','down','right','enter'])
            AP_blue_count += 1      #adds one to total blue count
        print(f'Blue APs: {AP_blue_count}')
    
    #def auto_script_total(self):
        #print(f'Total APs: ')
        #print(f'Total Failed APs:')

    def save_new_sheet(self):   #Pop-up message box, asks if user wants to save sorted date to new excell file/sheet
        result = messagebox.askyesno("Visio AP Coloring Tool","Do you want to save data in new Excell?")
        print(result)
        if result == True:
            print("data saved in 'New File'")
            
        else:
            pass


class AutoColor(tkinter.Frame):  #POP-UP GUI for choosing majority of Sheets color
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        
        #master = messagebox("Visio AP Coloring Tool\nSelect the AP color majority: ")
        self.pack()
        self.create_widgets()

    def create_widgets(self):   #buttons in the GUI
        #self.title.pack(text="""""", justify = Tk.LEFT, padx = 20)

        self.title = tkinter.Label(self, text = "Choose AP Color Majority: ")
        self.title.pack(side = "top")

        self.option1 = tkinter.Button(self, fg = "green")    #Button, colored green
        self.option1["text"] = "Green"                      #button, named green
        self.option1["command"] = self.output1              #when button pressed, execute output1
        self.option1.pack(side="top")                       #position selt at top most position

        self.option2 = tkinter.Button(self, text = "Red", fg = "red", command = self.output2)  #better formatted buttons
        self.option2.pack(side="top")

        self.option3 = tkinter.Button(self, text = "Orange", fg = "orange", command = self.output3)
        self.option3.pack(side="top")

        self.option4 = tkinter.Button(self, text = "Blue", fg = "blue", command = self.output4)
        self.option4.pack(side="top")

        self.quit = tkinter.Button(self, text="QUIT", fg="red", command=self.close)
        self.quit.pack(side="bottom", pady = 30)

    def output1(self):   #colors all but green
        print("GREEN AP Majority...")
        print("QUICK! You have 5 seconds to click into your visio file!!!\n")
        time.sleep(5)
        Options.auto_script_red(self)
        Options.auto_script_orange(self)
        Options.auto_script_blue(self)
        self.close()

    def output2(self):  #colors all but red
        print("RED AP Majority...")
        print("QUICK! You have 5 seconds to click into your visio file!!!\n")
        time.sleep(5)
        Options.auto_script_green(self)
        Options.auto_script_blue(self)
        Options.auto_script_orange(self)
        self.close()

    def output3(self):  #colors all but orange
        print("Orange AP Majority...")
        print("QUICK! You have 5 seconds to click into your visio file!!!\n")
        time.sleep(5)
        Options.auto_script_green(self)
        Options.auto_script_red(self)
        Options.auto_script_blue(self)
        self.close()

    def output4(self):  #colors all but blue
        print("Blue AP Majority...")
        print("QUICK! You have 5 seconds to click into your visio file!!!\n")
        time.sleep(5)
        Options.auto_script_green(self)
        Options.auto_script_red(self)
        Options.auto_script_orange(self)
        Options.save_new_sheet(self)
        self.close()
    
    def close(self):
        print("Closing program... Have a nice day   :)")
        time.sleep(1)
        exit(0)


def main():
    #choice = "" #creates a blank string for user input
    keyPress = ""
    #while choice != "3": 
    while keyPress != "3": 
        print("Enter '1' to open Visio AP Naming tool")
        print("Enter '2' to open Visio AP Coloring tool")
        print("Enter '3' to quit program\n")
        #choice = input(">>> ")
        keyPress = keyboard.read_key()
        if keyPress == "1":
            import VisioNameTool as vnt
            vnt.theGUI()    #the GUI function calls these internally:
            break               #vnt.startWorkbook()   
                                #vnt.openFiles()
                                #vnt.visioLoop()
                                    #vnt.visioGuts()
                                        #vnt.saveExcel()       
        elif keyPress == "2":
            #AutoColor Tool
            start = Options()
            start.open_audit_sheet()
            start.audit_sheet_sorter()
            start.search_options()
            root = tkinter.Tk()
            A_C = AutoColor(master=root)
            A_C.master.title("Visio AP Coloring Tool")
            A_C.mainloop()
            break
        
        else:
            print("Invalid input")
            pass

if __name__ == "__main__":
    main()   #this is the call to run the whole program
