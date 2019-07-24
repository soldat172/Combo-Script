import pyautogui, time, openpyxl, keyboard, datetime, sys, os
from openpyxl import Workbook
from tkinter import filedialog, Tk
from guizero import App, Text, TextBox, PushButton, Window

global siteName
global apPrefix
global apNumber
global apName
global excelCount
global excelNumber
global rowLetter
global rowNumber
global rowWholeName
global list
global excelName
global formatNumber
global excelStartingInteger
global keyPress

siteName = str ('Paul')
apPrefix = siteName + '-XX-'
apNumber =  int('1')
apName = str(apPrefix) + str(apNumber)
excelCount = int('1')
excelNumber = int('1')
rowLetter = 'A'
rowNumber = int('1')

rowWholeName = str(rowLetter) + str(rowNumber)
list=[99999]
excelName = 'Paul'
formatNumber = format(apNumber, '05')
excelStartingInteger = int('1')
currentTime = datetime.datetime.now().strftime("%Y-%m-%d  %H.%M.%S")

def startWorkbook ():#Starts an excel. Required for visioTool.
    global wb
    wb = Workbook()
    global ws
    ws = wb.active

def openFiles():#Opens file explorer
    global excelName
    #Tk().withdraw() # We don't want a full GUI, so keep the root window from appearing
    excelName = filedialog.askopenfilename() # Show an "Open" dialog box and return the path to the selected file
    print(excelName)
    print('')

def pause(): #Pause logic
    keyPress = keyboard.read_key()
    if keyPress == 'pause': #Reads if pause button has been pressed and unpauses
        print('Unpaused')
        time.sleep(.3)
        exit
    else: #If pause has not been pressed then loop repeats infinitely WOO
        pause()
        
def theGUI(): #all the GUI stuff
    print('')
    print('Start of Visio Tool')
    def apStartingNumber(): # part of GUI that allows changing of the starting AP number
        global apNumber
        global formatNumber
        apNumber = int(startingNumber.value)
        formatNumber = format(apNumber, '05')
        print ('AP number changed to ' + str(apNumber))
        changingText.value = "AP number changed to " + str(apNumber) #Text for changing AP number

    def changeSiteName(): # part of GUI that allows changing of the site name
        global siteName
        global apPrefix
        global apName
        siteName = str(siteNames.value)
        apPrefix = siteName + '-XX-'
        apName = str(apPrefix) + str(apNumber)
        print('Site name changed to ' + siteName)
        changingText.value = "Site changed to " + siteName #Text for changing site name
        
    def directions(): #Once "Go?" has been pressed execute the following items
        startWorkbook()
        openFiles()
        app.hide()
        visioLoop()
        python = sys.executable #Restarts the whole program
        os.execl(python, python, * sys.argv) #Restarts the whole program
        
    app = App(title = "Phoenix_Oath", width=352, height=132, layout='grid')

    button7 = PushButton(app, text = "Go?", command = directions, grid=[2,3])

    #Logic for changing starting AP number
    startingNumberText = Text(app, text="AP Number?", align="left", grid=[0,1]) #Text asks for AP number
    startingNumber = TextBox(app, align="right",text = "1", width=30, grid=[1,1]) #Text box for data entry
    button4 = PushButton(app, text = "Confirm", command = apStartingNumber, grid=[2,1])

    #Logic for changing site name
    siteNamesText = Text(app, text="Site Name?", align="left",  grid=[0,2]) #Text asks for site name
    siteNames = TextBox(app, align="right",text = "PAUL", width=30, grid=[1,2]) #Text box for data entry
    button6 = PushButton(app, text = "Confirm", command = changeSiteName, grid=[2,2])

    changingText = Text(app,text="War has changed", align ="left", grid=[1,3])
    app.display() # initiates the GUI. Allowing it to be used
    
def saveExcel(): # saves Visio Tool names to Excel
    global displayText
    global excelName
    wb.save(excelName) # Saves workbook
    print ('')
    print ('Worksheet saved')
    print ('End of Visio Tool')
        
def visioGuts(): # the internals to the Visio Tool. Determines how most of the program is run
    global apNumber
    global formatNumber
    global excelStartingInteger
    print (apName)
    ws.cell(excelStartingInteger, 1, apName)  #writes in excel **format** ->(row, column, content to be written in cell)
    pyautogui.press('backspace'); pyautogui.typewrite(str(formatNumber)) # takes control of keyboard. hits backspace and types AP number
    apNumber += 1 #increments ap number up by 1
    excelStartingInteger += 1 # Increments Excel cell to be written in
    formatNumber = format(apNumber, '05') # Modifies apNumber by adding up to 5 zeros in front

def visioLoop():
    global apNumber
    global formatNumber
    global excelStartingInteger
    global apName
    while True:
        keyPress = keyboard.read_key() #reads ALL keypress's and saves to variable
        
        if keyPress == '`' : #Adds a Default Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber)
            visioGuts ()
            
        elif keyPress == 's': #Adds a Standup Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber) + 'S'
            visioGuts ()
            
        elif keyPress == 'g' : #Adds a Guard Shack Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber) + 'G'
            visioGuts ()
            
        elif keyPress == 'm' : #Adds a MOD/KIVA Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber) + 'M'
            visioGuts ()
            
        elif keyPress == 'h' : #Adds a High Racking Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber) + 'H'
            visioGuts ()
            
        elif keyPress == 'd' : #Adds a Door Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber) + 'D'
            visioGuts ()

        elif keyPress == 'e' : #Adds a External Ap to cutsheet
            apName = str(apPrefix) + str(formatNumber) + 'E'
            visioGuts ()
            
        elif keyPress == '[': #goes down one ap number.
            apNumber -= 1
            excelStartingInteger -= 1
            formatNumber = format(apNumber, '05')
            apName = str(apPrefix) + str(formatNumber)
            print (apName + " fix?")
            time.sleep(.09)

        elif keyPress == ']': #goes up one ap number.
            apNumber += 1
            excelStartingInteger += 1
            formatNumber = format(apNumber, '05')
            apName = str(apPrefix) + str(formatNumber)
            print (apName + " fix?")
            time.sleep(.09)
            
        elif keyPress == 'pause': #Well it pauses everthing.....
            print('Paused')
            time.sleep(.3)
            pause() #Program is stuck in the pause loop until pause is pressed again

        elif keyPress == '=': #Saves the CAD Cutsheet and ends the Visio Tool   
            saveExcel()
            exit(0)

