import pyautogui, time, openpyxl, keyboard, datetime, sys, os
from openpyxl import Workbook
from tkinter import filedialog, Tk
from guizero import App, Text, TextBox, PushButton, Window

#-------------------------------------------------------------------
global siteName
global apPrefix
global apNumber
global apName
global excelCount
global excelNumber
global rowLetter
global rowNumber
global rowWholeName
global list
global excelName
global formatNumber
global excelStartingInteger
global keyPress

siteName = str ('Paul')
apPrefix = siteName + '-XX-'
apNumber =  int('1')
apName = str(apPrefix) + str(apNumber)
excelCount = int('1')
excelNumber = int('1')
rowLetter = 'A'
rowNumber = int('1')
rowWholeName = str(rowLetter) + str(rowNumber)
list=[99999]
excelName = 'Paul'
formatNumber = format(apNumber, '05')
excelStartingInteger = int('1')
currentTime = datetime.datetime.now().strftime("%Y-%m-%d  %H.%M.%S")

baseX=1920 #The base resolution at which the sample "clicks" were made from
baseY=1200 #The base resolution at which the sample "clicks" were made from
newX,newY = pyautogui.size() #Finds current monitor resolution

#-------------------------------------------------------------------

if baseX != newX:
    multiplierX = 1.2786 #Hard Code
    multiplierY = 1.1535 #Hard Code
    multiplierX2 = 1.4481 #Hard Code
    multiplierY2 = 1.1012 #Hard Code
else:
    multiplierX = newX/baseX #Finds how mucher bigger/smaller the "x" resolution is
    multiplierY = newY/baseY #Finds how mucher bigger/smaller the "y" resolution is
    multiplierX2 = newX/baseX #Hard Code
    multiplierY2 = newY/baseY #Hard Code

#The location of the button is multiplied by the change in resolution 
#so that this program works on any resolution
nextButtonX = 1152 * multiplierX #X Location of the "next" button while using DataExtraction
nextButtonY = 775 * multiplierY #Y Location of the "next" button while using DataExtraction
desktopButtonX = 703 * multiplierX2 #X Location of the "desktop icon" button while using DataExtraction
desktopButtonY = 652 * multiplierY2 #Y Location of the "desktop icon" button while using DataExtraction
saveButtonX = 1231 * multiplierX #X Location of the "save" button while using DataExtraction
saveButtonY = 775 * multiplierY #Y Location of the "save" button while using DataExtraction

#-------------------------------------------------------------------

def startWorkbook ():#Starts an excel. Required for visioTool.
    global wb
    wb = Workbook()
    global ws
    ws = wb.active
    #visioLoop()

def openFiles():#Opens file explorer
    global excelName
    Tk().withdraw() # We don't want a full GUI, so keep the root window from appearing
    excelName = filedialog.askopenfilename() # Show an "Open" dialog box and return the path to the selected file
    print(excelName)
    print('')

def pause(): #Pause logic
    keyPress = keyboard.read_key()
    if keyPress == 'pause': #Reads if pause button has been pressed and unpauses
        print('unpaused')
        time.sleep(.3)
        exit
    else: #If pause has not been pressed then loop repeats infinitely WOO
        #time.sleep(.3)
        pause()
        
def theGUI(): #all the GUI stuff
    def apStartingNumber(): # part of GUI that allows changing of the starting AP number
        global apNumber
        global formatNumber
        apNumber = int(startingNumber.value)
        formatNumber = format(apNumber, '05')
        print ('AP number changed to ' + str(apNumber))
        changingText.value = "AP number changed to " + str(apNumber) #Text for changing AP number

    def changeSiteName(): # part of GUI that allows changing of the site name
        global siteName
        global apPrefix
        global apName
        siteName = str(siteNames.value)
        apPrefix = siteName + '-XX-'
        apName = str(apPrefix) + str(apNumber)
        print('Site name changed to ' + siteName)
        changingText.value = "Site changed to " + siteName #Text for changing site name

    def repeat ():
        theCreator()
        theGUI()
        print ('please')
        
    def directions():
        startWorkbook()
        openFiles()
        app.hide()
        visioLoop()
        repeat()
        print ('hmm')
        
    app = App(title = "Phoenix_Oath", width=352, height=132, layout='grid')

    button7 = PushButton(app, text = "Go?", command = directions, grid=[2,3])

    #Logic for changing starting AP number
    startingNumberText = Text(app, text="AP Number?", align="left", grid=[0,1]) #Text asks for AP number
    startingNumber = TextBox(app, align="right",text = "1", width=30, grid=[1,1]) #Text box for data entry
    button4 = PushButton(app, text = "Confirm", command = apStartingNumber, grid=[2,1])

    #Logic for changing site name
    siteNamesText = Text(app, text="Site Name?", align="left",  grid=[0,2]) #Text asks for site name
    siteNames = TextBox(app, align="right",text = "PAUL", width=30, grid=[1,2]) #Text box for data entry
    button6 = PushButton(app, text = "Confirm", command = changeSiteName, grid=[2,2])

    changingText = Text(app,text="War has changed", align ="left", grid=[1,3])
    app.display() # initiates the GUI. Allowing it to be used
    
def saveExcel(): # saves Visio Tool names to Excel
    global displayText
    global excelName
    wb.save(excelName) # Saves workbook
    print ('End of Visio Tool')
        
def visioGuts(): # the internals to the Visio Tool. Determines how most of the program is run
    global apNumber
    global formatNumber
    global excelStartingInteger
    print (apName)
    ws.cell(excelStartingInteger, 1, apName)  #writes in excel **format** ->(row, column, content to be written in cell)
    pyautogui.press('backspace'); pyautogui.press('backspace')
    pyautogui.typewrite(str(formatNumber)) # takes control of keyboard. hits backspace and types AP number
    apNumber += 1 #increments ap number up by 1
    excelStartingInteger += 1 # Increments Excel cell to be written in
    formatNumber = format(apNumber, '05') # Modifies apNumber by adding up to 5 zeros in front

def visioLoop():
    global apNumber
    global formatNumber
    global excelStartingInteger
    global apName
    while True:
        keyPress = keyboard.read_key() #reads ALL keypress's and saves to variable
        
        if keyPress == '`' :
            apName = str(apPrefix) + str(formatNumber)
            visioGuts ()
            
        elif keyPress == 's':
            apName = str(apPrefix) + str(formatNumber) + 'S'
            visioGuts ()
            
        elif keyPress == 'g' :
            apName = str(apPrefix) + str(formatNumber) + 'G'
            visioGuts ()
            
        elif keyPress == 'm' :
            apName = str(apPrefix) + str(formatNumber) + 'M'
            visioGuts ()
            
        elif keyPress == 'h' :
            apName = str(apPrefix) + str(formatNumber) + 'H'
            visioGuts ()
            
        elif keyPress == 'd' :
            apName = str(apPrefix) + str(formatNumber) + 'D'
            visioGuts ()

        elif keyPress == 'e' :
            apName = str(apPrefix) + str(formatNumber) + 'E'
            visioGuts ()
            
        elif keyPress == '[':  #goes down one ap number.
            apNumber -= 1
            excelStartingInteger -= 1
            formatNumber = format(apNumber, '05')
            apName = str(apPrefix) + str(formatNumber)
            print (apName)

        elif keyPress == ']':  #goes up one ap number.    
            apNumber += 1
            excelStartingInteger += 1
            formatNumber = format(apNumber, '05')
            apName = str(apPrefix) + str(formatNumber)
            print (apName)
            
        elif keyPress == 'pause': #Well it pauses or unpauses everthing.....
            print('paused')
            time.sleep(.3)
            pause()


        elif keyPress == '=': #ends the Visio Tool   
            saveExcel()
            print('test')
            break

def theCreator():
    actionDelay = .5 #How long of a pause before a action is done.
    clickDelay = .3 #How long of a pause before a click is done
    while True: 
        keyPress = keyboard.read_key() #reads ALL keypress's and saves to variable
        currentTime = datetime.datetime.now().strftime("%Y-%m-%d  %H.%M.%S")
        
        #Saves Visio drawing as a CAD drawing
        if keyPress == '[' : #What key needs to be pressed to activate the "if" statement
            pyautogui.hotkey('alt', 'f2',interval=.1)
            pyautogui.moveTo(500, 100) #Moves mouse in order to avoid highlighting and messing up the proccess
            pyautogui.typewrite ('Coord ' + (currentTime),interval=.1) #Name of file saved
            pyautogui.hotkey('alt', 't',interval=.1)
            pyautogui.hotkey('alt', 'down',interval=.1)
            pyautogui.typewrite(['down','down','down','down','down','down','down','down','down','down','down','down','down'], interval=.1)
            pyautogui.typewrite(['enter','enter'], interval=.1)

        #Scales up CAD to drawing units
        if keyPress == ']' : #What key needs to be pressed to activate the "if" statement
            pyautogui.press('esc',interval=.1)
            pyautogui.hotkey('ctrlleft', 'a',interval=.1)
            pyautogui.typewrite ('scale',interval=.1)
            pyautogui.press('enter',interval=.1)
            pyautogui.typewrite(['0',',','0',',','0'], interval=.1)
            pyautogui.press('enter',interval=.1)
            pyautogui.typewrite('12', interval=.1) #How much to scale up
            pyautogui.press('enter',interval=.1)
            pyautogui.press ('esc',interval=.5) #Data extraction portion, from here down
            pyautogui.typewrite ('dataextraction',interval=.02)
            pyautogui.press ('enter',interval=actionDelay)
            pyautogui.click(x= nextButtonX , y= nextButtonY ,interval=clickDelay)
            pyautogui.typewrite ('Coord ' + (currentTime),interval=.05) #Name of data Extraction
            pyautogui.click(x= desktopButtonX , y= desktopButtonY ,interval=clickDelay)
            pyautogui.click(x= saveButtonX , y= saveButtonY ,interval=clickDelay)
            pyautogui.click(x= nextButtonX , y= nextButtonY ,interval=clickDelay)
            time.sleep(3) #long delay for the loading screen
            pyautogui.press (['tab','tab'],interval=actionDelay) 
            pyautogui.hotkey ('ctrlleft','a',interval=.1)
            pyautogui.press ('space',interval=actionDelay)
            pyautogui.press ('up',interval=actionDelay)
            pyautogui.press (['down','down'],interval=actionDelay)
            pyautogui.press ('space',interval=actionDelay)
            pyautogui.click(x= nextButtonX , y= nextButtonY ,interval=clickDelay)
            pyautogui.press (['tab','tab'],interval=actionDelay)
            pyautogui.hotkey ('ctrlleft','a',interval=.05)
            pyautogui.press ('space',interval=actionDelay)
            pyautogui.press ('up',interval=actionDelay)
            pyautogui.press (['down','down','down'],interval=.05)
            pyautogui.press ('space',interval=actionDelay)
            pyautogui.press (['down','down','down','down','down','down','down','down','down','down','down','down','down','down','down','down','down','down','down','down','down'],interval=.01)
            pyautogui.press ('space',interval=actionDelay)
            pyautogui.press ('down',interval=actionDelay)
            pyautogui.press ('space',interval=actionDelay)
            pyautogui.click(x= nextButtonX , y= nextButtonY ,clicks=2,interval=clickDelay)
            pyautogui.press (['tab','tab'],interval=.1)
            pyautogui.press ('space',interval=actionDelay)
            pyautogui.press ('tab',interval=actionDelay)
            pyautogui.press ('space',interval=actionDelay)
            pyautogui.typewrite ('Coord ' + (currentTime),interval=.05) #Name of data Extraction
            pyautogui.click(x= desktopButtonX , y= desktopButtonY ,interval=clickDelay)
            pyautogui.click(x= saveButtonX , y= saveButtonY ,interval=clickDelay)
            pyautogui.click(x= nextButtonX , y= nextButtonY ,clicks=2,interval=clickDelay)

        elif keyPress == '\\':
            break       

        elif keyPress == 'pause': #Well it pauses or unpauses everthing.....
            print('paused')
            time.sleep(.3)
            pause()
            
theCreator() #First call of program
theGUI()
print('should not be here')
